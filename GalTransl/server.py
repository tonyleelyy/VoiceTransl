from __future__ import annotations

import argparse
import json
import threading
from concurrent.futures import ThreadPoolExecutor
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import urlparse, parse_qs, unquote
from uuid import uuid4

import os
import re
from datetime import datetime
from collections import deque
from dataclasses import asdict, dataclass, field
from yaml import safe_load, safe_dump
from packaging.version import InvalidVersion, Version

from base64 import urlsafe_b64decode, urlsafe_b64encode

from GalTransl import TRANSLATOR_SUPPORTED, INPUT_FOLDERNAME, OUTPUT_FOLDERNAME, CACHE_FOLDERNAME, GALTRANSL_VERSION, new_version
from GalTransl.Service import JobSpec, JobState, create_job_state, run_job
from GalTransl.AppSettings import load_app_settings, save_app_settings
from GalTransl.DefaultProjectConfig import DEFAULT_PROJECT_CONFIG_YAML


def _utcnow_text() -> str:
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def _normalize_version_text(value: str) -> str:
    return value.strip().removeprefix("v").removeprefix("V")


def _has_newer_release(current_version: str, latest_version: str | None) -> bool:
    if not latest_version:
        return False

    current_text = _normalize_version_text(current_version)
    latest_text = _normalize_version_text(latest_version)
    try:
        return Version(latest_text) > Version(current_text)
    except InvalidVersion:
        return latest_text != current_text


def _normalize_project_dir(project_dir: str) -> str:
    return str(Path(project_dir).resolve())


class _ConcurrentLimitError(ValueError):
    """Raised when the global concurrent job limit has been reached."""
    pass


@dataclass(slots=True)
class RuntimeSentenceEvent:
    id: str
    ts: str
    filename: str
    index: int
    speaker: str | list[str] | None
    source_preview: str
    translation_preview: str
    trans_by: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(slots=True)
class RuntimeErrorEvent:
    id: str
    ts: str
    kind: str
    level: str
    message: str
    filename: str = ""
    index_range: str = ""
    retry_count: int | None = None
    model: str = ""
    sleep_seconds: float | None = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


RUNTIME_RECENT_EVENT_LIMIT = 80
RUNTIME_PER_FILE_SUCCESS_LIMIT = 100
# Upper bound on the flat list of success events returned per snapshot. Each
# translating file keeps its own 100-slot deque, but returning all of them every
# poll would quickly explode the HTTP payload. 500 is enough to satisfy the
# UI's 100-card render budget plus any per-file filter on a small number of
# concurrently active files.
RUNTIME_SNAPSHOT_SUCCESS_LIMIT = 500


@dataclass(slots=True)
class RuntimeState:
    project_dir: str
    workers_active: int = 0
    workers_configured: int = 0
    stage: str = ""
    current_file: str = ""
    updated_at: str = field(default_factory=_utcnow_text)
    file_totals: dict[str, int] = field(default_factory=dict)
    cache_file_display_map: dict[str, str] = field(default_factory=dict)
    # Per-file deque of recent success events. Each file keeps up to
    # RUNTIME_PER_FILE_SUCCESS_LIMIT events independently so that concurrent
    # translations of multiple files do not evict each other's cards.
    # Each deque stores newest-first (appendleft) for O(1) merging.
    recent_successes_by_file: dict[str, deque[RuntimeSentenceEvent]] = field(default_factory=dict)
    recent_errors: deque[RuntimeErrorEvent] = field(default_factory=lambda: deque(maxlen=RUNTIME_RECENT_EVENT_LIMIT))
    success_timestamps: deque[float] = field(default_factory=deque)


class RuntimeRegistry:
    def __init__(self) -> None:
        self._states: dict[str, RuntimeState] = {}
        self._lock = threading.Lock()

    def ensure_project(self, project_dir: str) -> RuntimeState:
        normalized = _normalize_project_dir(project_dir)
        with self._lock:
            state = self._states.get(normalized)
            if state is None:
                state = RuntimeState(project_dir=project_dir)
                self._states[normalized] = state
            else:
                state.project_dir = project_dir
            state.updated_at = _utcnow_text()
            return state

    def reset_project(self, project_dir: str) -> None:
        with self._lock:
            normalized = _normalize_project_dir(project_dir)
            self._states[normalized] = RuntimeState(project_dir=project_dir)

    def update_status(
        self,
        project_dir: str,
        *,
        stage: str | None = None,
        current_file: str | None = None,
        workers_active: int | None = None,
        workers_configured: int | None = None,
        file_totals: dict[str, int] | None = None,
        cache_file_display_map: dict[str, str] | None = None,
    ) -> None:
        with self._lock:
            state = self._states.get(_normalize_project_dir(project_dir))
            if state is None:
                state = RuntimeState(project_dir=project_dir)
                self._states[_normalize_project_dir(project_dir)] = state
            if stage is not None:
                state.stage = stage
            if current_file is not None:
                state.current_file = current_file
            if workers_active is not None:
                state.workers_active = max(0, workers_active)
            if workers_configured is not None:
                state.workers_configured = max(0, workers_configured)
            if file_totals is not None:
                state.file_totals = dict(file_totals)
            if cache_file_display_map is not None:
                state.cache_file_display_map = dict(cache_file_display_map)
            state.updated_at = _utcnow_text()

    def append_success(
        self,
        project_dir: str,
        *,
        filename: str,
        index: int,
        speaker: str | list[str] | None,
        source_preview: str,
        translation_preview: str,
        trans_by: str = "",
    ) -> None:
        now = datetime.utcnow().timestamp()
        with self._lock:
            state = self._states.get(_normalize_project_dir(project_dir))
            if state is None:
                state = RuntimeState(project_dir=project_dir)
                self._states[_normalize_project_dir(project_dir)] = state
            display_filename = self._resolve_display_filename_locked(state, filename)
            event = RuntimeSentenceEvent(
                id=f"{display_filename}:{index}:{int(now * 1000)}",
                ts=_utcnow_text(),
                filename=display_filename,
                index=index,
                speaker=speaker,
                source_preview=_trim_preview(source_preview),
                translation_preview=_trim_preview(translation_preview),
                trans_by=trans_by,
            )
            file_deque = state.recent_successes_by_file.get(display_filename)
            if file_deque is None:
                file_deque = deque(maxlen=RUNTIME_PER_FILE_SUCCESS_LIMIT)
                state.recent_successes_by_file[display_filename] = file_deque
            file_deque.appendleft(event)
            state.success_timestamps.append(now)
            self._trim_speed_window_locked(state, now)
            state.updated_at = event.ts

    def append_error(
        self,
        project_dir: str,
        *,
        kind: str,
        message: str,
        filename: str = "",
        index_range: str = "",
        retry_count: int | None = None,
        model: str = "",
        sleep_seconds: float | None = None,
        level: str = "error",
    ) -> None:
        with self._lock:
            state = self._states.get(_normalize_project_dir(project_dir))
            if state is None:
                state = RuntimeState(project_dir=project_dir)
                self._states[_normalize_project_dir(project_dir)] = state
            display_filename = self._resolve_display_filename_locked(state, filename)
            ts = _utcnow_text()
            state.recent_errors.appendleft(RuntimeErrorEvent(
                id=f"{kind}:{display_filename}:{int(datetime.utcnow().timestamp() * 1000)}",
                ts=ts,
                kind=kind,
                level=level,
                message=_trim_preview(message, 240),
                filename=display_filename,
                index_range=index_range,
                retry_count=retry_count,
                model=model,
                sleep_seconds=sleep_seconds,
            ))
            state.updated_at = ts

    @staticmethod
    def _resolve_display_filename_locked(state: RuntimeState, filename: str) -> str:
        normalized = str(filename or "").strip()
        if not normalized:
            return ""

        candidates: list[str] = []

        def add_candidate(value: str) -> None:
            candidate = str(value or "").strip()
            if candidate and candidate not in candidates:
                candidates.append(candidate)

        add_candidate(normalized)
        add_candidate(f"{normalized}.json")
        add_candidate(f"{normalized}{_CACHE_APPEND_SUFFIX}")

        split_match = re.match(r"^(.*)_\d+$", normalized)
        if split_match:
            split_base = split_match.group(1)
            add_candidate(split_base)
            add_candidate(f"{split_base}.json")
            add_candidate(f"{split_base}{_CACHE_APPEND_SUFFIX}")

        normalized_path = normalized.replace("-}", "/")
        add_candidate(normalized_path)
        add_candidate(f"{normalized_path}.json")
        add_candidate(f"{normalized_path}{_CACHE_APPEND_SUFFIX}")

        split_path_match = re.match(r"^(.*)_\d+$", normalized_path)
        if split_path_match:
            split_path_base = split_path_match.group(1)
            add_candidate(split_path_base)
            add_candidate(f"{split_path_base}.json")
            add_candidate(f"{split_path_base}{_CACHE_APPEND_SUFFIX}")

        for candidate in candidates:
            display = state.cache_file_display_map.get(candidate)
            if display:
                return display

        if normalized_path in state.file_totals:
            return normalized_path
        if split_path_match and split_path_match.group(1) in state.file_totals:
            return split_path_match.group(1)
        if normalized in state.file_totals:
            return normalized

        return normalized_path

    def get_runtime_snapshot(self, project_dir: str) -> dict[str, Any]:
        normalized = _normalize_project_dir(project_dir)
        with self._lock:
            state = self._states.get(normalized)
            if state is None:
                return {
                    "stage": "",
                    "current_file": "",
                    "workers_active": 0,
                    "workers_configured": 0,
                    "translation_speed_lpm": 0,
                    "file_totals": {},
                    "cache_file_display_map": {},
                    "recent_errors": [],
                    "recent_successes": [],
                    "updated_at": _utcnow_text(),
                }
            now = datetime.utcnow().timestamp()
            self._trim_speed_window_locked(state, now)
            speed = round((len(state.success_timestamps) / 60) * 60, 1) if state.success_timestamps else 0
            # Flatten per-file success deques (each newest-first) and re-order
            # globally by timestamp desc so the snapshot list remains newest-first
            # for existing clients.
            merged_successes: list[RuntimeSentenceEvent] = []
            for file_deque in state.recent_successes_by_file.values():
                merged_successes.extend(file_deque)
            merged_successes.sort(key=lambda ev: ev.ts, reverse=True)
            if len(merged_successes) > RUNTIME_SNAPSHOT_SUCCESS_LIMIT:
                merged_successes = merged_successes[:RUNTIME_SNAPSHOT_SUCCESS_LIMIT]
            return {
                "stage": state.stage,
                "current_file": state.current_file,
                "workers_active": state.workers_active,
                "workers_configured": state.workers_configured,
                "translation_speed_lpm": speed,
                "file_totals": dict(state.file_totals),
                "cache_file_display_map": dict(state.cache_file_display_map),
                "recent_errors": [event.to_dict() for event in state.recent_errors],
                "recent_successes": [event.to_dict() for event in merged_successes],
                "updated_at": state.updated_at,
            }

    @staticmethod
    def _trim_speed_window_locked(state: RuntimeState, now: float) -> None:
        while state.success_timestamps and now - state.success_timestamps[0] > 60:
            state.success_timestamps.popleft()


def _trim_preview(value: str, limit: int = 140) -> str:
    normalized = str(value or "").replace("\r", " ").replace("\n", " ").strip()
    if len(normalized) <= limit:
        return normalized
    return normalized[: max(0, limit - 1)] + "…"


RUNTIME_REGISTRY = RuntimeRegistry()
_CACHE_APPEND_SUFFIX = ".append.jsonl"


@dataclass(slots=True)
class _CacheProgressFileStat:
    mtime_ns: int
    size: int
    translated_keys: frozenset[str]
    problem_keys: frozenset[str]
    failed_keys: frozenset[str]
    retran_terms_signature: tuple[str, ...] = field(default_factory=tuple)
    retran_hit_keys: dict[str, frozenset[str]] = field(default_factory=dict)


@dataclass(slots=True)
class _RetranConfigStat:
    mtime_ns: int
    size: int
    retran_key: str | list[str]


def _normalize_retran_key(value: Any) -> str | list[str]:
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return [str(item) for item in value if item]
    return ""


def _normalize_retran_terms(value: str | list[str]) -> list[str]:
    if isinstance(value, str):
        normalized = value.strip()
        return [normalized] if normalized else []
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            text = str(item or "").strip()
            if text:
                result.append(text)
        return result
    return []


def _check_retran_key(retran_key: str | list[str], target: Any) -> bool:
    text = str(target or "")
    if isinstance(retran_key, str):
        return bool(retran_key) and retran_key in text
    if isinstance(retran_key, list):
        return any(key in text for key in retran_key if key)
    return False


def _parse_runtime_job_started_at_ns(value: str) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        normalized = text.replace("Z", "+00:00")
        dt = datetime.fromisoformat(normalized)
        return int(dt.timestamp() * 1_000_000_000)
    except Exception:
        return None


class RuntimeProgressCache:
    def __init__(self) -> None:
        self._project_files: dict[str, dict[str, _CacheProgressFileStat]] = {}
        self._retran_config_cache: dict[str, _RetranConfigStat] = {}
        self._lock = threading.Lock()

    def reset_project(self, project_dir: str) -> None:
        normalized = _normalize_project_dir(project_dir)
        with self._lock:
            self._project_files.pop(normalized, None)

    def get_retran_key(self, project_dir: str, config_file_name: str = "config.yaml") -> str | list[str]:
        config_path = os.path.join(project_dir, config_file_name or "config.yaml")
        normalized_config = str(Path(config_path).resolve())

        try:
            stat = os.stat(config_path)
        except OSError:
            with self._lock:
                self._retran_config_cache.pop(normalized_config, None)
            return ""

        with self._lock:
            cached = self._retran_config_cache.get(normalized_config)
            if (
                cached is not None
                and cached.mtime_ns == int(stat.st_mtime_ns)
                and cached.size == int(stat.st_size)
            ):
                return cached.retran_key

        retran_key: str | list[str] = ""
        try:
            with open(config_path, "rb") as cfg_file:
                cfg = safe_load(cfg_file.read()) or {}
            common = cfg.get("common", {}) if isinstance(cfg, dict) else {}
            retran_key = _normalize_retran_key(common.get("retranslKey", ""))
        except Exception:
            retran_key = ""

        with self._lock:
            self._retran_config_cache[normalized_config] = _RetranConfigStat(
                mtime_ns=int(stat.st_mtime_ns),
                size=int(stat.st_size),
                retran_key=retran_key,
            )

        return retran_key

    def get_progress(
        self,
        project_dir: str,
        file_totals: dict[str, int],
        cache_file_display_map: dict[str, str],
        retran_key: str | list[str] = "",
        retran_terms: list[str] | None = None,
        current_job_started_at_ns: int | None = None,
    ) -> dict[str, Any]:
        normalized = _normalize_project_dir(project_dir)
        cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
        retran_terms = retran_terms or []
        retran_terms_signature = tuple(retran_terms)

        with self._lock:
            project_stats = self._project_files.setdefault(normalized, {})
            seen_files: set[str] = set()

            if os.path.isdir(cache_dir):
                for entry in os.scandir(cache_dir):
                    if not entry.is_file():
                        continue
                    if not (
                        entry.name.endswith(".json")
                        or entry.name.endswith(_CACHE_APPEND_SUFFIX)
                    ):
                        continue

                    name = entry.name
                    seen_files.add(name)

                    try:
                        stat = entry.stat()
                    except OSError:
                        continue

                    cached = project_stats.get(name)
                    if (
                        cached is not None
                        and cached.mtime_ns == int(stat.st_mtime_ns)
                        and cached.size == int(stat.st_size)
                        and cached.retran_terms_signature == retran_terms_signature
                    ):
                        continue

                    translated_keys: set[str] = set()
                    problem_keys: set[str] = set()
                    failed_keys: set[str] = set()
                    retran_hit_keys: dict[str, set[str]] = {term: set() for term in retran_terms}

                    def _name_src(items: list[Any], idx: int) -> str:
                        if idx < 0 or idx >= len(items):
                            return ""
                        item = items[idx]
                        if not isinstance(item, dict):
                            return ""
                        name = str(item.get("name", "") or "")
                        pre_src = str(item.get("pre_src", item.get("pre_jp", "")) or "")
                        return f"{name}{pre_src}"

                    def _entry_signature(items: list[Any], idx: int) -> str:
                        line_now = _name_src(items, idx)
                        row = items[idx] if 0 <= idx < len(items) else {}
                        row_index = str(row.get("index", "")) if isinstance(row, dict) else ""
                        if not line_now:
                            if isinstance(row, dict):
                                row_src = str(
                                    row.get("pre_src", row.get("pre_jp", row.get("post_src", "")))
                                    or ""
                                )
                                row_name = str(row.get("name", "") or "")
                                return f"__row__:{idx}:{row_index}:{row_name}:{row_src}"
                            return f"__row__:{idx}"

                        line_prev = "None"
                        j = idx - 1
                        while j >= 0:
                            candidate = _name_src(items, j)
                            if candidate:
                                line_prev = candidate
                                break
                            j -= 1

                        line_next = "None"
                        j = idx + 1
                        while j < len(items):
                            candidate = _name_src(items, j)
                            if candidate:
                                line_next = candidate
                                break
                            j += 1

                        # 在 context key 前拼接 entry 的 index，使相同上下文三元组但位于
                        # 不同位置的条目（如重复短句）生成不同 key，避免 set 去重导致
                        # 进度少计。index 前缀同时保证 .json 与 .append.jsonl 对同一
                        # 位置条目仍能正确去重（二者 index 相同 → key 相同）。
                        context_key = f"{line_prev}{line_now}{line_next}"
                        if row_index:
                            return f"{row_index}:{context_key}"
                        return context_key

                    entries: list[Any] = []
                    try:
                        import orjson
                        with open(entry.path, "rb") as f:
                            raw = f.read()
                        if entry.name.endswith(_CACHE_APPEND_SUFFIX):
                            for line in raw.splitlines():
                                if not line:
                                    continue
                                try:
                                    row = orjson.loads(line)
                                except Exception:
                                    continue
                                if isinstance(row, dict):
                                    entries.append(row)
                        else:
                            loaded = orjson.loads(raw)
                            if isinstance(loaded, list):
                                entries = loaded
                    except Exception:
                        continue

                    for idx, item in enumerate(entries):
                        if not isinstance(item, dict):
                            continue

                        entry_key = str(item.get("__cache_key", "")).strip()
                        if entry_key:
                            # 同 _entry_signature：以 index 为前缀使不同位置的同文本条目
                            # 在 set 中各占一席，同时保持与 .json 快照 key 的一致性。
                            item_index = str(item.get("index", ""))
                            if item_index:
                                entry_key = f"{item_index}:{entry_key}"
                        else:
                            entry_key = _entry_signature(entries, idx)

                        is_translated = bool(item.get("pre_dst", "") or item.get("pre_zh", ""))
                        is_problem = bool(item.get("problem", ""))
                        is_failed = (
                            "翻译失败" in str(item.get("problem", ""))
                            or "(Failed)" in str(item.get("pre_dst", "") or item.get("pre_zh", ""))
                            or "(翻译失败)" in str(item.get("pre_dst", "") or item.get("pre_zh", ""))
                        )

                        no_proofread = str(item.get("proofread_dst", "") or "") == ""
                        should_apply_retransl_filter = not entry.name.endswith(_CACHE_APPEND_SUFFIX)
                        if (
                            should_apply_retransl_filter
                            and current_job_started_at_ns is not None
                            and int(stat.st_mtime_ns) >= int(current_job_started_at_ns)
                        ):
                            should_apply_retransl_filter = False
                        if (
                            is_translated
                            and should_apply_retransl_filter
                            and retran_key
                            and no_proofread
                            and (
                                _check_retran_key(retran_key, item.get("pre_src", item.get("pre_jp", "")))
                                or _check_retran_key(retran_key, item.get("problem", ""))
                            )
                        ):
                            is_translated = False

                        if (
                            is_translated
                            and not entry.name.endswith(_CACHE_APPEND_SUFFIX)
                            and no_proofread
                            and retran_hit_keys
                        ):
                            source_text = item.get("pre_src", item.get("pre_jp", ""))
                            problem_text = item.get("problem", "")
                            for term in retran_terms:
                                if _check_retran_key(term, source_text) or _check_retran_key(term, problem_text):
                                    retran_hit_keys[term].add(entry_key)

                        if is_translated:
                            translated_keys.add(entry_key)
                        if is_problem:
                            problem_keys.add(entry_key)
                        if is_failed:
                            failed_keys.add(entry_key)

                    project_stats[name] = _CacheProgressFileStat(
                        mtime_ns=int(stat.st_mtime_ns),
                        size=int(stat.st_size),
                        translated_keys=frozenset(translated_keys),
                        problem_keys=frozenset(problem_keys),
                        failed_keys=frozenset(failed_keys),
                        retran_terms_signature=retran_terms_signature,
                        retran_hit_keys={
                            term: frozenset(hit_keys)
                            for term, hit_keys in retran_hit_keys.items()
                        },
                    )

            stale_files = [name for name in project_stats if name not in seen_files]
            for name in stale_files:
                project_stats.pop(name, None)

            file_progress_map: dict[str, dict[str, Any]] = {}
            retran_counts: dict[str, set[str]] = {term: set() for term in retran_terms}

            for name, stat in project_stats.items():
                canonical_name = (
                    name[: -len(_CACHE_APPEND_SUFFIX)]
                    if name.endswith(_CACHE_APPEND_SUFFIX)
                    else name
                )
                display_name = cache_file_display_map.get(canonical_name, canonical_name)
                if file_totals and display_name not in file_totals:
                    continue
                if display_name not in file_progress_map:
                    file_progress_map[display_name] = {
                        "filename": display_name,
                        "total": int(file_totals.get(display_name, 0)),
                        "translated": 0,
                        "problems": 0,
                        "failed": 0,
                        "_translated_keys": set(),
                        "_problem_keys": set(),
                        "_failed_keys": set(),
                    }
                file_progress_map[display_name]["_translated_keys"].update(stat.translated_keys)
                file_progress_map[display_name]["_problem_keys"].update(stat.problem_keys)
                file_progress_map[display_name]["_failed_keys"].update(stat.failed_keys)
                for term, hit_keys in stat.retran_hit_keys.items():
                    retran_counts.setdefault(term, set()).update(hit_keys)

            for display_name, total_count in file_totals.items():
                file_progress_map.setdefault(
                    display_name,
                    {
                        "filename": display_name,
                        "total": int(total_count),
                        "translated": 0,
                        "problems": 0,
                        "failed": 0,
                        "_translated_keys": set(),
                        "_problem_keys": set(),
                        "_failed_keys": set(),
                    },
                )

            for file_progress in file_progress_map.values():
                total_count = int(file_progress.get("total", 0))
                translated = len(file_progress["_translated_keys"])
                problems = len(file_progress["_problem_keys"])
                failed = len(file_progress["_failed_keys"])

                if total_count > 0:
                    translated = min(translated, total_count)
                    problems = min(problems, total_count)
                    failed = min(failed, total_count)

                file_progress["translated"] = translated
                file_progress["problems"] = problems
                file_progress["failed"] = failed
                file_progress.pop("_translated_keys", None)
                file_progress.pop("_problem_keys", None)
                file_progress.pop("_failed_keys", None)

            files = sorted(file_progress_map.values(), key=lambda item: item["filename"])
            return {
                "total": sum(int(item["total"]) for item in files),
                "translated": sum(int(item["translated"]) for item in files),
                "problems": sum(int(item["problems"]) for item in files),
                "failed": sum(int(item["failed"]) for item in files),
                "retransl_stats": [
                    {"key": term, "count": len(retran_counts.get(term, set()))}
                    for term in retran_terms
                ],
                "files": files,
            }


RUNTIME_PROGRESS_CACHE = RuntimeProgressCache()


def reset_runtime_project(project_dir: str) -> None:
    RUNTIME_REGISTRY.reset_project(project_dir)
    RUNTIME_PROGRESS_CACHE.reset_project(project_dir)


def update_runtime_status(
    project_dir: str,
    *,
    stage: str | None = None,
    current_file: str | None = None,
    workers_active: int | None = None,
    workers_configured: int | None = None,
    file_totals: dict[str, int] | None = None,
    cache_file_display_map: dict[str, str] | None = None,
) -> None:
    RUNTIME_REGISTRY.update_status(
        project_dir,
        stage=stage,
        current_file=current_file,
        workers_active=workers_active,
        workers_configured=workers_configured,
        file_totals=file_totals,
        cache_file_display_map=cache_file_display_map,
    )


def record_runtime_success(
    project_dir: str,
    *,
    filename: str,
    index: int,
    speaker: str | list[str] | None,
    source_preview: str,
    translation_preview: str,
    trans_by: str = "",
) -> None:
    RUNTIME_REGISTRY.append_success(
        project_dir,
        filename=filename,
        index=index,
        speaker=speaker,
        source_preview=source_preview,
        translation_preview=translation_preview,
        trans_by=trans_by,
    )


def record_runtime_error(
    project_dir: str,
    *,
    kind: str,
    message: str,
    filename: str = "",
    index_range: str = "",
    retry_count: int | None = None,
    model: str = "",
    sleep_seconds: float | None = None,
    level: str = "error",
) -> None:
    RUNTIME_REGISTRY.append_error(
        project_dir,
        kind=kind,
        message=message,
        filename=filename,
        index_range=index_range,
        retry_count=retry_count,
        model=model,
        sleep_seconds=sleep_seconds,
        level=level,
    )


# ---------------------------------------------------------------------------
# Project path helpers – encode/decode directory paths for use in URLs
# ---------------------------------------------------------------------------

def encode_project_dir(project_dir: str) -> str:
    """Encode a filesystem path to a URL-safe token."""
    return urlsafe_b64encode(project_dir.encode("utf-8")).decode("ascii")


def decode_project_dir(token: str) -> str:
    """Decode a URL-safe token back to a filesystem path."""
    # Restore base64 padding that may have been stripped for URL safety
    padding = 4 - len(token) % 4
    if padding != 4:
        token += "=" * padding
    return urlsafe_b64decode(token.encode("ascii")).decode("utf-8")


# ---------------------------------------------------------------------------
# Filesystem helpers for project API
# ---------------------------------------------------------------------------

def _safe_project_dir(token: str) -> str:
    """Decode and validate a project directory token. Raises ValueError on failure."""
    try:
        project_dir = decode_project_dir(token)
    except Exception:
        raise ValueError("invalid project id")
    if not os.path.isdir(project_dir):
        raise ValueError(f"project directory does not exist: {project_dir}")
    return project_dir


def _read_yaml_file(path: str) -> dict:
    """Read and parse a YAML file."""
    with open(path, "r", encoding="utf-8") as f:
        return safe_load(f) or {}


def _write_yaml_file(path: str, data: dict) -> None:
    """Write data to a YAML file atomically."""
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        safe_dump(data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
    os.replace(tmp, path)


def _list_dir_entries(dir_path: str, *, count_json_entries: bool = False) -> list[dict[str, Any]]:
    """List files in a directory with basic metadata."""
    entries = []
    if not os.path.isdir(dir_path):
        return entries
    for name in sorted(os.listdir(dir_path)):
        full = os.path.join(dir_path, name)
        stat = os.stat(full) if os.path.isfile(full) else None
        entry = {
            "name": name,
            "is_file": os.path.isfile(full),
            "size": stat.st_size if stat else 0,
            "modified": datetime.fromtimestamp(stat.st_mtime).isoformat() if stat else "",
        }
        if count_json_entries and name.endswith(".json") and os.path.isfile(full):
            try:
                with open(full, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    entry["entry_count"] = len(data)
            except Exception:
                pass
        entries.append(entry)
    return entries


DICT_PROJECT_MARKER = "(project_dir)"
COMMON_DICT_CATEGORY_MAP = ".category_map.json"


def _is_safe_dict_filename(filename: str) -> bool:
    if not isinstance(filename, str):
        return False
    trimmed = filename.strip()
    if not trimmed:
        return False
    if trimmed != os.path.basename(trimmed):
        return False
    if any(sep in trimmed for sep in ("/", "\\")):
        return False
    return True


def _is_safe_config_filename(filename: str) -> bool:
    if not isinstance(filename, str):
        return False
    trimmed = filename.strip()
    if not trimmed:
        return False
    if trimmed != os.path.basename(trimmed):
        return False
    if any(sep in trimmed for sep in ("/", "\\")):
        return False
    if ".." in trimmed:
        return False
    return True


def _is_path_within(base_dir: str, target_path: str) -> bool:
    base_abs = os.path.abspath(base_dir)
    target_abs = os.path.abspath(target_path)
    try:
        common = os.path.commonpath([base_abs, target_abs])
    except ValueError:
        return False
    return common == base_abs


def _normalize_dict_text(content: str) -> str:
    return str(content or "").replace("\r\n", "\n").replace("\r", "\n")


def _dict_category_config_key(category: str) -> str:
    if category == "pre":
        return "preDict"
    if category == "gpt":
        return "gpt.dict"
    if category == "post":
        return "postDict"
    raise ValueError(f"invalid dictionary category: {category}")


def _read_dict_file_payload(file_path: str) -> dict[str, Any]:
    if not os.path.isfile(file_path):
        return {
            "path": file_path,
            "lines": [],
            "count": 0,
            "mtime": None,
            "error": "file not found",
        }
    try:
        mtime = os.path.getmtime(file_path)
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.read().splitlines()
        return {
            "path": file_path,
            "lines": lines,
            "count": len([
                line_item
                for line_item in lines
                if line_item.strip() and not line_item.startswith("\\\\") and not line_item.startswith("//")
            ]),
            "mtime": mtime,
        }
    except Exception:
        return {
            "path": file_path,
            "lines": [],
            "count": 0,
            "mtime": None,
            "error": "failed to read",
        }


def _collect_project_dict_payload(project_dir: str, config_name: str) -> dict[str, Any]:
    if not _is_safe_config_filename(config_name):
        raise ValueError("invalid config filename")
    config_path = os.path.join(project_dir, config_name)
    if not os.path.isfile(config_path):
        raise FileNotFoundError(f"config file not found: {config_name}")

    data = _read_yaml_file(config_path)
    dict_cfg = data.get("dictionary", {})

    pre_all = [str(x) for x in dict_cfg.get("preDict", [])]
    gpt_all = [str(x) for x in dict_cfg.get("gpt.dict", [])]
    post_all = [str(x) for x in dict_cfg.get("postDict", [])]

    pre_files = [x for x in pre_all if x.startswith(DICT_PROJECT_MARKER)]
    gpt_files = [x for x in gpt_all if x.startswith(DICT_PROJECT_MARKER)]
    post_files = [x for x in post_all if x.startswith(DICT_PROJECT_MARKER)]

    dict_contents: dict[str, dict[str, Any]] = {}
    for file_key in pre_files + gpt_files + post_files:
        clean = file_key.replace(DICT_PROJECT_MARKER, "").strip()
        if not _is_safe_dict_filename(clean):
            dict_contents[file_key] = {
                "path": os.path.join(project_dir, clean),
                "lines": [],
                "count": 0,
                "mtime": None,
                "error": "invalid dictionary filename",
            }
            continue
        file_path = os.path.join(project_dir, clean)
        if not _is_path_within(project_dir, file_path):
            dict_contents[file_key] = {
                "path": file_path,
                "lines": [],
                "count": 0,
                "mtime": None,
                "error": "dictionary path escapes project directory",
            }
            continue
        dict_contents[file_key] = _read_dict_file_payload(file_path)

    return {
        "project_dir": project_dir,
        "config_file_name": config_name,
        "pre_dict_files": pre_files,
        "gpt_dict_files": gpt_files,
        "post_dict_files": post_files,
        "dict_contents": dict_contents,
    }


def _common_dict_directory() -> str:
    return os.path.abspath("Dict")


def _ensure_project_dict_file_configured(project_dir: str, config_name: str, category: str, filename: str) -> None:
    if not _is_safe_config_filename(config_name):
        raise ValueError("invalid config filename")
    if not _is_safe_dict_filename(filename):
        raise ValueError("invalid dictionary filename")

    config_path = os.path.join(project_dir, config_name)
    if not os.path.isfile(config_path):
        raise FileNotFoundError(f"config file not found: {config_name}")

    data = _read_yaml_file(config_path)
    dict_cfg_raw = data.get("dictionary", {})
    dict_cfg = dict_cfg_raw if isinstance(dict_cfg_raw, dict) else {}
    list_key = _dict_category_config_key(category)
    current_items = dict_cfg.get(list_key, [])

    if isinstance(current_items, list):
        current_list = [str(x) for x in current_items]
    elif current_items in (None, ""):
        current_list = []
    else:
        current_list = [str(current_items)]

    file_key = f"{DICT_PROJECT_MARKER}{filename}"
    if file_key in current_list:
        return

    current_list.append(file_key)
    dict_cfg[list_key] = current_list
    data["dictionary"] = dict_cfg
    _write_yaml_file(config_path, data)


def _common_dict_category_map_path(dict_dir: str) -> str:
    return os.path.join(dict_dir, COMMON_DICT_CATEGORY_MAP)


def _read_common_dict_category_map(dict_dir: str) -> dict[str, str]:
    map_path = _common_dict_category_map_path(dict_dir)
    if not os.path.isfile(map_path):
        return {}
    try:
        with open(map_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {}
        result: dict[str, str] = {}
        for key, value in data.items():
            if _is_safe_dict_filename(str(key)) and str(value) in {"pre", "gpt", "post"}:
                result[str(key)] = str(value)
        return result
    except Exception:
        return {}


def _write_common_dict_category_map(dict_dir: str, category_map: dict[str, str]) -> None:
    map_path = _common_dict_category_map_path(dict_dir)
    tmp_path = map_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(category_map, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, map_path)


def _categorize_common_dict_file(filename: str) -> str:
    lower = filename.lower()
    if "gpt" in lower:
        return "gpt"
    if "post" in lower or "译后" in filename:
        return "post"
    return "pre"


def _collect_common_dict_payload() -> dict[str, Any]:
    dict_dir = _common_dict_directory()
    os.makedirs(dict_dir, exist_ok=True)
    category_map = _read_common_dict_category_map(dict_dir)

    files = [
        name
        for name in sorted(os.listdir(dict_dir))
        if os.path.isfile(os.path.join(dict_dir, name)) and name != COMMON_DICT_CATEGORY_MAP
    ]

    pre_files: list[str] = []
    gpt_files: list[str] = []
    post_files: list[str] = []
    dict_contents: dict[str, dict[str, Any]] = {}

    for name in files:
        category = category_map.get(name) or _categorize_common_dict_file(name)
        if category == "gpt":
            gpt_files.append(name)
        elif category == "post":
            post_files.append(name)
        else:
            pre_files.append(name)
        dict_contents[name] = _read_dict_file_payload(os.path.join(dict_dir, name))

    return {
        "dict_dir": dict_dir,
        "pre_dict_files": pre_files,
        "gpt_dict_files": gpt_files,
        "post_dict_files": post_files,
        "dict_contents": dict_contents,
    }


# ---------------------------------------------------------------------------
# Global backend profiles helpers
# ---------------------------------------------------------------------------

_BACKEND_PROFILES_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "backend_profiles.yaml")


def _read_backend_profiles() -> dict:
    """Read the global backend profiles YAML file."""
    if not os.path.isfile(_BACKEND_PROFILES_PATH):
        return {"profiles": {}}
    try:
        return _read_yaml_file(_BACKEND_PROFILES_PATH)
    except Exception:
        return {"profiles": {}}


def _write_backend_profiles(data: dict) -> None:
    """Write the global backend profiles YAML file atomically."""
    _write_yaml_file(_BACKEND_PROFILES_PATH, data)


_PROBLEM_TYPE_CATALOG: list[dict[str, str]] = [
    {"name": "词频过高", "description": "某字在译文中重复大于 20 次（且远多于原文）。"},
    {"name": "标点错漏", "description": "括号/引号/冒号等标点与原文不一致。"},
    {"name": "残留日文", "description": "译文中残留日文平假名或片假名。"},
    {"name": "丢失换行", "description": "译文缺少原文中的行内换行。"},
    {"name": "多加换行", "description": "译文换行符比原文多，可能导致溢出。"},
    {"name": "比日文长", "description": "译文长度超过原文 1.3 倍（常用，宽松阈值）。"},
    {"name": "比日文长严格", "description": "译文长度超过原文（零容忍，严格阈值）。"},
    {"name": "字典使用", "description": "没有按 GPT 字典的要求翻译。"},
    {"name": "引入英文", "description": "原文无英文，但译文引入了英文单词。"},
    {"name": "语言不通", "description": "译文包含大量非 GBK 字符（仅对中文目标语言生效）。"},
    {"name": "缺控制符", "description": "译文缺少原文中的控制符（如 \\n、变量标记等）。"},
    {"name": "独白男他", "description": "独白（无name）译文出现'他'。"},
]


def _list_problem_types() -> list[dict[str, str]]:
    """Return the list of problem types supported by the backend analyzer.

    Each entry has ``name`` (used in YAML config) and a short ``description``.
    Kept in sync with :class:`GalTransl.ConfigHelper.CProblemType` and
    :func:`GalTransl.Problem.find_problems`.
    """
    return list(_PROBLEM_TYPE_CATALOG)


def _list_translation_guidelines() -> list[str]:
    """List translation guideline filenames under the ``translation_guidelines`` folder."""
    guidelines_dir = os.path.abspath("translation_guidelines")
    if not os.path.isdir(guidelines_dir):
        return []
    result: list[str] = []
    for name in sorted(os.listdir(guidelines_dir)):
        full = os.path.join(guidelines_dir, name)
        if not os.path.isfile(full):
            continue
        lower = name.lower()
        if lower.endswith(".md") or lower.endswith(".txt"):
            result.append(name)
    return result


def _scan_plugins() -> list[dict[str, Any]]:
    """Scan the plugins directory and return plugin metadata."""
    plugins_dir = os.path.abspath("plugins")
    result = []
    if not os.path.isdir(plugins_dir):
        return result
    for name in sorted(os.listdir(plugins_dir)):
        yaml_path = os.path.join(plugins_dir, name, f"{name}.yaml")
        if not os.path.isfile(yaml_path):
            continue
        try:
            info = _read_yaml_file(yaml_path)
            core = info.get("Core", {})
            settings = info.get("Settings", {})
            result.append({
                "name": name,
                "display_name": core.get("Name", name),
                "version": core.get("Version", ""),
                "author": core.get("Author", ""),
                "description": core.get("Description", ""),
                "type": core.get("Type", "unknown").lower(),
                "module": core.get("Module", name),
                "settings": settings,
            })
        except Exception:
            continue
    return result


INDEX_HTML = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>GalTransl Backend Mode</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f5f7fb;
      --panel: #ffffff;
      --line: #d9e1ec;
      --text: #1f2a37;
      --muted: #5b6b7f;
      --primary: #2f6feb;
      --primary-hover: #1d5fe0;
      --success: #0f9d58;
      --warning: #f39c12;
      --danger: #d93025;
    }

    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Segoe UI", "Microsoft YaHei", sans-serif;
      background: var(--bg);
      color: var(--text);
    }
    .page {
      max-width: 1080px;
      margin: 0 auto;
      padding: 24px;
    }
    .hero {
      margin-bottom: 20px;
    }
    .hero h1 {
      margin: 0 0 8px;
      font-size: 30px;
    }
    .hero p {
      margin: 0;
      color: var(--muted);
      line-height: 1.6;
    }
    .layout {
      display: grid;
      grid-template-columns: minmax(320px, 420px) 1fr;
      gap: 20px;
      align-items: start;
    }
    .card {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 16px;
      padding: 18px;
      box-shadow: 0 8px 30px rgba(16, 24, 40, 0.05);
    }
    .card h2 {
      margin: 0 0 16px;
      font-size: 18px;
    }
    .field {
      margin-bottom: 14px;
    }
    .field label {
      display: block;
      font-size: 13px;
      color: var(--muted);
      margin-bottom: 6px;
    }
    .field input,
    .field select {
      width: 100%;
      padding: 10px 12px;
      border: 1px solid var(--line);
      border-radius: 10px;
      font-size: 14px;
      background: #fff;
    }
    .actions {
      display: flex;
      gap: 10px;
      align-items: center;
      margin-top: 10px;
    }
    button {
      border: 0;
      background: var(--primary);
      color: white;
      padding: 10px 16px;
      border-radius: 10px;
      font-size: 14px;
      cursor: pointer;
    }
    button:hover { background: var(--primary-hover); }
    button.secondary {
      background: #e8eef9;
      color: var(--primary);
    }
    .hint {
      font-size: 13px;
      color: var(--muted);
      line-height: 1.6;
      margin-top: 12px;
    }
    .pill {
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      padding: 4px 10px;
      font-size: 12px;
      font-weight: 600;
    }
    .status-pending { background: #eef3ff; color: #3159c9; }
    .status-running { background: #fff4dd; color: #a96500; }
    .status-completed { background: #e7f8ee; color: var(--success); }
    .status-failed { background: #fdecea; color: var(--danger); }
    .status-cancelled { background: #f1f3f5; color: #667085; }
    .job-list {
      display: grid;
      gap: 12px;
    }
    .job {
      border: 1px solid var(--line);
      border-radius: 14px;
      padding: 14px;
      background: #fff;
    }
    .job-header {
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
      margin-bottom: 8px;
    }
    .job-title {
      font-weight: 600;
      overflow-wrap: anywhere;
    }
    .job-meta {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 8px 14px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.5;
      margin-top: 10px;
    }
    .job-error {
      margin-top: 10px;
      color: var(--danger);
      background: #fff4f4;
      border-radius: 10px;
      padding: 10px 12px;
      font-size: 13px;
      white-space: pre-wrap;
    }
    .empty {
      color: var(--muted);
      font-size: 14px;
      padding: 18px 0;
    }
    @media (max-width: 860px) {
      .layout { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <div class="page">
    <section class="hero">
      <h1>GalTransl Backend Mode</h1>
      <p>这是一个为后续 Web UI 铺路的最小验证界面。它不会替换原来的 <code>run_GalTransl.py</code>，而是通过独立的本地服务入口提交和查看翻译任务。</p>
    </section>

    <div class="layout">
      <section class="card">
        <h2>提交任务</h2>
        <form id="job-form">
          <div class="field">
            <label for="project_dir">项目目录</label>
            <input id="project_dir" name="project_dir" placeholder="例如：E:\\GalTransl\\sampleProject" required />
          </div>
          <div class="field">
            <label for="config_file_name">配置文件名</label>
            <input id="config_file_name" name="config_file_name" value="config.yaml" required />
          </div>
          <div class="field">
            <label for="translator">翻译模板</label>
            <select id="translator" name="translator" required></select>
          </div>
          <div class="actions">
            <button type="submit">启动任务</button>
            <button type="button" class="secondary" id="refresh-btn">刷新状态</button>
          </div>
        </form>
        <div class="hint">
          建议先用 <code>show-plugs</code> 或现有可工作的项目配置来验证服务链路。当前版本采用单执行槽，避免和共享日志、缓存输出发生冲突。
        </div>
      </section>

      <section class="card">
        <h2>任务列表</h2>
        <div id="jobs" class="job-list"></div>
      </section>
    </div>
  </div>

  <script>
    const translatorSelect = document.getElementById('translator');
    const jobsContainer = document.getElementById('jobs');
    const form = document.getElementById('job-form');
    const refreshBtn = document.getElementById('refresh-btn');

    function escapeHtml(value) {
      return String(value ?? '')
        .replaceAll('&', '&amp;')
        .replaceAll('<', '&lt;')
        .replaceAll('>', '&gt;')
        .replaceAll('"', '&quot;')
        .replaceAll("'", '&#39;');
    }

    function statusClass(status) {
      return `status-${status || 'pending'}`;
    }

    async function loadTranslators() {
      const response = await fetch('/api/translators');
      const data = await response.json();
      translatorSelect.innerHTML = data.translators.map(item => {
        return `<option value="${escapeHtml(item.name)}">${escapeHtml(item.name)} - ${escapeHtml(item.description)}</option>`;
      }).join('');
    }

    function renderJobs(jobs) {
      if (!jobs.length) {
        jobsContainer.innerHTML = '<div class="empty">还没有任务，先从左侧提交一个本地任务。</div>';
        return;
      }

      jobsContainer.innerHTML = jobs.map(job => `
        <article class="job">
          <div class="job-header">
            <div class="job-title">${escapeHtml(job.project_dir)}</div>
            <span class="pill ${statusClass(job.status)}">${escapeHtml(job.status)}</span>
          </div>
          <div>${escapeHtml(job.translator)} / ${escapeHtml(job.config_file_name)}</div>
          <div class="job-meta">
            <div><strong>任务 ID：</strong>${escapeHtml(job.job_id)}</div>
            <div><strong>创建时间：</strong>${escapeHtml(job.created_at || '-')}</div>
            <div><strong>开始时间：</strong>${escapeHtml(job.started_at || '-')}</div>
            <div><strong>结束时间：</strong>${escapeHtml(job.finished_at || '-')}</div>
            <div><strong>执行结果：</strong>${job.success ? 'success' : 'not finished / failed'}</div>
          </div>
          ${job.error ? `<div class="job-error">${escapeHtml(job.error)}</div>` : ''}
        </article>
      `).join('');
    }

    async function loadJobs() {
      const response = await fetch('/api/jobs');
      const data = await response.json();
      renderJobs(data.jobs || []);
    }

    form.addEventListener('submit', async (event) => {
      event.preventDefault();
      const payload = {
        project_dir: document.getElementById('project_dir').value,
        config_file_name: document.getElementById('config_file_name').value,
        translator: translatorSelect.value,
      };

      const response = await fetch('/api/jobs', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(payload),
      });
      const data = await response.json();

      if (!response.ok) {
        alert(data.error || '提交任务失败');
        return;
      }

      await loadJobs();
    });

    refreshBtn.addEventListener('click', () => loadJobs());
    loadTranslators().then(loadJobs);
    setInterval(loadJobs, 2000);
  </script>
</body>
</html>
"""


class JobRegistry:
    def __init__(self, max_workers: int | None = None) -> None:
        self._jobs: dict[str, JobState] = {}
        self._stop_events: dict[str, threading.Event] = {}
        self._lock = threading.Lock()
        self._max_workers = max_workers or load_app_settings().get("maxConcurrentJobs", 4)
        self._executor = ThreadPoolExecutor(max_workers=self._max_workers, thread_name_prefix="galtransl-job")

    def list_jobs(self) -> list[dict[str, Any]]:
        with self._lock:
            jobs = [job.to_dict() for job in self._jobs.values()]
        return sorted(jobs, key=lambda job: job["created_at"], reverse=True)

    def get_job(self, job_id: str) -> dict[str, Any] | None:
        with self._lock:
            job = self._jobs.get(job_id)
            return None if job is None else job.to_dict()

    def get_project_job(self, project_dir: str) -> JobState | None:
        normalized = _normalize_project_dir(project_dir)
        with self._lock:
            project_jobs = [
                job for job in self._jobs.values()
                if _normalize_project_dir(job.project_dir) == normalized
            ]
        if not project_jobs:
            return None
        active_jobs = [job for job in project_jobs if job.status in {"pending", "running"}]
        if active_jobs:
            active_jobs.sort(key=lambda job: job.created_at, reverse=True)
            return active_jobs[0]
        project_jobs.sort(key=lambda job: job.created_at, reverse=True)
        return project_jobs[0]

    def request_project_stop(self, project_dir: str) -> JobState | None:
        normalized = _normalize_project_dir(project_dir)
        with self._lock:
            active_jobs = [
                job for job in self._jobs.values()
                if _normalize_project_dir(job.project_dir) == normalized and job.status in {"pending", "running"}
            ]
            if not active_jobs:
                return None
            active_jobs.sort(key=lambda job: job.created_at, reverse=True)
            event = self._stop_events.get(normalized)
            if event is not None:
                event.set()
            return active_jobs[0]

    def clear_project_stop(self, project_dir: str) -> None:
        normalized = _normalize_project_dir(project_dir)
        with self._lock:
            self._stop_events.pop(normalized, None)

    def _has_running_job_for_project(self, project_dir: str) -> bool:
        normalized = str(Path(project_dir).resolve())
        for job in self._jobs.values():
            if str(Path(job.project_dir).resolve()) == normalized and job.status in {"pending", "running"}:
                return True
        return False

    def _running_job_count(self) -> int:
        return sum(1 for job in self._jobs.values() if job.status in {"pending", "running"})

    def submit(self, payload: dict[str, Any]) -> dict[str, Any]:
        project_dir = str(payload.get("project_dir", "")).strip()
        config_file_name = str(payload.get("config_file_name", "config.yaml")).strip() or "config.yaml"
        translator = str(payload.get("translator", "")).strip()
        backend_profile = str(payload.get("backend_profile", "")).strip()
        backend_profile_data = payload.get("backend_profile_data")

        if not project_dir:
            raise ValueError("project_dir is required")
        if not translator:
            raise ValueError("translator is required")
        if translator not in TRANSLATOR_SUPPORTED:
            raise ValueError(f"unsupported translator: {translator}")
        if translator == "GenDic":
            _ensure_project_dict_file_configured(
                project_dir,
                config_file_name,
                "gpt",
                "项目GPT字典-生成.txt",
            )

        with self._lock:
            if self._has_running_job_for_project(project_dir):
                raise ValueError("the project already has a pending or running job")
            if self._running_job_count() >= self._max_workers:
                raise _ConcurrentLimitError(f"已达到最大并发翻译任务数 ({self._max_workers})，请等待已有任务完成后再试")

            job_id = uuid4().hex[:12]
            spec = JobSpec(
                job_id=job_id,
                project_dir=project_dir,
                config_file_name=config_file_name,
                translator=translator,
                backend_profile=backend_profile,
                backend_profile_data=backend_profile_data if isinstance(backend_profile_data, dict) else {},
            )
            state = create_job_state(spec)
            reset_runtime_project(project_dir)
            self._jobs[job_id] = state
            self._stop_events[_normalize_project_dir(project_dir)] = threading.Event()
            self._executor.submit(self._execute_job, spec, state)
            return state.to_dict()

    def _execute_job(self, spec: JobSpec, state: JobState) -> None:
        stop_event = self._stop_events.get(_normalize_project_dir(spec.project_dir))
        try:
            run_job(spec, state, stop_event=stop_event)
        finally:
            self.clear_project_stop(spec.project_dir)


def build_handler(registry: JobRegistry):
    class RequestHandler(BaseHTTPRequestHandler):
        def end_headers(self) -> None:
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type")
            super().end_headers()

        def do_OPTIONS(self) -> None:
            self.send_response(HTTPStatus.NO_CONTENT)
            self.end_headers()

        # -----------------------------------------------------------------
        # Routing helpers
        # -----------------------------------------------------------------

        def _route_project_api(self, project_id: str, sub_path: str) -> None:
            """Handle /api/projects/:id/* routes."""
            try:
                project_dir = _safe_project_dir(project_id)
            except ValueError as exc:
                self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
                return

            # GET /api/projects/:id/config
            if sub_path == "/config":
                config_name = parse_qs(urlparse(self.path).query).get("config", ["config.yaml"])[0]
                config_path = os.path.join(project_dir, config_name)
                if not os.path.isfile(config_path):
                    self._send_json({"error": f"config file not found: {config_name}"}, status=HTTPStatus.NOT_FOUND)
                    return
                try:
                    data = _read_yaml_file(config_path)
                    self._send_json({"config": data, "project_dir": project_dir, "config_file_name": config_name})
                except Exception as exc:
                    self._send_json({"error": f"failed to read config: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # GET /api/projects/:id/files
            if sub_path == "/files":
                input_dir = os.path.join(project_dir, INPUT_FOLDERNAME)
                output_dir = os.path.join(project_dir, OUTPUT_FOLDERNAME)
                cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                self._send_json({
                    "project_dir": project_dir,
                    "input_dir": input_dir,
                    "output_dir": output_dir,
                    "cache_dir": cache_dir,
                    "input_files": _list_dir_entries(input_dir),
                    "output_files": _list_dir_entries(output_dir),
                    "cache_files": _list_dir_entries(cache_dir, count_json_entries=True),
                })
                return

            # GET /api/projects/:id/cache
            if sub_path == "/cache":
                cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                self._send_json({
                    "project_dir": project_dir,
                    "cache_dir": cache_dir,
                    "files": _list_dir_entries(cache_dir, count_json_entries=True),
                })
                return

            # POST /api/projects/:id/cache/save
            if sub_path == "/cache/save":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    payload = self._read_json_body()
                    filename = str(payload.get("filename", "")).strip()
                    entries = payload.get("entries", [])
                    config_name = str(payload.get("config_file_name", "config.yaml")).strip() or "config.yaml"

                    if not filename or filename != os.path.basename(filename):
                        self._send_json({"error": "invalid cache filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                    file_path = os.path.join(cache_dir, filename)
                    if not os.path.isfile(file_path):
                        self._send_json({"error": f"cache file not found: {filename}"}, status=HTTPStatus.NOT_FOUND)
                        return

                    import orjson
                    with open(file_path, "wb") as f:
                        f.write(orjson.dumps(entries, option=orjson.OPT_INDENT_2))

                    # Rebuild: re-derive problem and post_dst_preview fields
                    try:
                        from GalTransl.CSentense import CSentense
                        from GalTransl.Problem import find_problems
                        from GalTransl.Frontend.LLMTranslate import preprocess_trans_list, postprocess_trans_list

                        # Load project config and dictionaries for rebuild
                        proj_config = None
                        pre_dic = None
                        post_dic = None
                        gpt_dic = None
                        tPlugins = []
                        try:
                            from GalTransl.ConfigHelper import CProjectConfig, initDictList
                            from GalTransl.Dictionary import CNormalDic, CGptDict
                            proj_config = CProjectConfig(project_dir, config_name)
                            dict_cfg = proj_config.getDictCfgSection()
                            pre_dic_list = dict_cfg.get("preDict", [])
                            post_dic_list = dict_cfg.get("postDict", [])
                            gpt_dic_list = dict_cfg.get("gpt.dict", [])
                            default_dic_dir = dict_cfg.get("defaultDictFolder", "")
                            pre_dic = CNormalDic(
                                initDictList(pre_dic_list, default_dic_dir, project_dir)
                            )
                            post_dic = CNormalDic(
                                initDictList(post_dic_list, default_dic_dir, project_dir)
                            )
                            gpt_dic = CGptDict(
                                initDictList(gpt_dic_list, default_dic_dir, project_dir)
                            )
                            if dict_cfg.get("sortDict", True):
                                pre_dic.sort_dic()
                                post_dic.sort_dic()
                                gpt_dic.sort_dic()
                            # Load text plugins
                            try:
                                tPlugins = proj_config.tPlugins
                            except Exception:
                                tPlugins = []
                        except Exception:
                            pass  # If config loading fails, skip dict processing

                        # Build CSentense list from saved entries
                        trans_list = []
                        for e in entries:
                            speaker = e.get("name", "")
                            if isinstance(speaker, list):
                                speaker = "/".join(speaker)
                            pre_src = e.get("pre_src", "") or e.get("pre_jp", "")
                            post_src = e.get("post_src", "") or e.get("post_jp", "")
                            pre_dst = e.get("pre_dst", "") or e.get("pre_zh", "")
                            proofread_dst = e.get("proofread_dst", "") or e.get("proofread_zh", "")
                            if post_src == "":
                                continue
                            s = CSentense(pre_src, speaker if speaker else "", e.get("index", 0))
                            s.post_jp = pre_src
                            s.pre_zh = pre_dst
                            s.proofread_zh = proofread_dst
                            s.post_zh = proofread_dst if proofread_dst else pre_dst
                            s.trans_by = e.get("trans_by", "")
                            s.proofread_by = e.get("proofread_by", "")
                            s.trans_conf = e.get("trans_conf", 0)
                            s.doub_content = e.get("doub_content", "")
                            s.unknown_proper_noun = e.get("unknown_proper_noun", "")
                            trans_list.append(s)

                        # Link prev/next
                        for i, s in enumerate(trans_list):
                            if i > 0:
                                s.prev_tran = trans_list[i - 1]
                            if i < len(trans_list) - 1:
                                s.next_tran = trans_list[i + 1]

                        # Pre-processing and post-processing (shared with LLMTranslate)
                        if pre_dic and proj_config:
                            preprocess_trans_list(trans_list, proj_config, pre_dic, tPlugins or None)
                        if post_dic and proj_config:
                            postprocess_trans_list(trans_list, proj_config, post_dic, tPlugins or None)

                        # Run find_problems
                        if trans_list:
                            try:
                                find_problems(trans_list, proj_config, gpt_dic)
                            except Exception:
                                pass  # If problem detection fails, skip

                        # Update entries with problem and post_dst_preview
                        idx = 0
                        for e in entries:
                            post_src_val = e.get("post_src", "") or e.get("post_jp", "")
                            if post_src_val == "":
                                continue
                            if idx < len(trans_list):
                                tran = trans_list[idx]
                                if tran.problem:
                                    e["problem"] = tran.problem
                                elif "problem" in e:
                                    del e["problem"]
                                e["post_dst_preview"] = tran.post_zh
                                idx += 1

                        # Re-save with updated fields
                        with open(file_path, "wb") as f:
                            f.write(orjson.dumps(entries, option=orjson.OPT_INDENT_2))

                        self._send_json({"success": True, "filename": filename, "entries": entries})
                    except Exception:
                        # Rebuild failed, but original save succeeded
                        self._send_json({"success": True, "filename": filename})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to save cache file: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/projects/:id/cache/delete-entry
            if sub_path == "/cache/delete-entry":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    payload = self._read_json_body()
                    filename = str(payload.get("filename", "")).strip()
                    entry_index = int(payload.get("index", -1))

                    if not filename or filename != os.path.basename(filename):
                        self._send_json({"error": "invalid cache filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                    file_path = os.path.join(cache_dir, filename)
                    if not os.path.isfile(file_path):
                        self._send_json({"error": f"cache file not found: {filename}"}, status=HTTPStatus.NOT_FOUND)
                        return

                    import orjson
                    with open(file_path, "rb") as f:
                        data = orjson.loads(f.read())
                    if not isinstance(data, list) or entry_index < 0 or entry_index >= len(data):
                        self._send_json({"error": "invalid entry index"}, status=HTTPStatus.BAD_REQUEST)
                        return
                    deleted = data.pop(entry_index)
                    with open(file_path, "wb") as f:
                        f.write(orjson.dumps(data, option=orjson.OPT_INDENT_2))
                    self._send_json({"success": True, "filename": filename, "deleted_index": entry_index})
                except (json.JSONDecodeError, ValueError):
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to delete cache entry: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/projects/:id/cache/delete-file
            if sub_path == "/cache/delete-file":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    payload = self._read_json_body()
                    filenames = payload.get("filenames", [])
                    if not isinstance(filenames, list) or not filenames:
                        self._send_json({"error": "filenames must be a non-empty list"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                    deleted_files = []
                    not_found_files = []
                    for fname in filenames:
                        fname = str(fname).strip()
                        if not fname or fname != os.path.basename(fname):
                            not_found_files.append(fname)
                            continue
                        file_path = os.path.join(cache_dir, fname)
                        if not os.path.isfile(file_path):
                            not_found_files.append(fname)
                            continue
                        try:
                            os.remove(file_path)
                            deleted_files.append(fname)
                        except OSError:
                            not_found_files.append(fname)
                    self._send_json({"success": True, "deleted_files": deleted_files, "not_found_files": not_found_files})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to delete cache files: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/projects/:id/cache/search
            if sub_path == "/cache/search":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    payload = self._read_json_body()
                    query = str(payload.get("query", "")).strip()
                    field = str(payload.get("field", "all")).strip()  # all | src | dst
                    max_results = min(int(payload.get("max_results", 500)), 2000)

                    if not query:
                        self._send_json({"results": [], "total": 0})
                        return

                    cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                    results = []
                    if os.path.isdir(cache_dir):
                        for name in sorted(os.listdir(cache_dir)):
                            if not name.endswith(".json"):
                                continue
                            fp = os.path.join(cache_dir, name)
                            if not os.path.isfile(fp):
                                continue
                            try:
                                import orjson
                                with open(fp, "rb") as f:
                                    entries = orjson.loads(f.read())
                                for e in entries:
                                    if not isinstance(e, dict):
                                        continue
                                    src_text = e.get("post_src", "") or e.get("post_jp", "") or e.get("pre_src", "") or e.get("pre_jp", "")
                                    dst_text = e.get("pre_dst", "") or e.get("pre_zh", "") or e.get("proofread_dst", "") or e.get("proofread_zh", "")
                                    problem_text = e.get("problem", "")
                                    match_src = query.lower() in src_text.lower()
                                    match_dst = query.lower() in dst_text.lower()
                                    match_problem = query.lower() in problem_text.lower()
                                    if field == "src" and not match_src:
                                        continue
                                    if field == "dst" and not match_dst:
                                        continue
                                    if field == "problem" and not match_problem:
                                        continue
                                    if field == "all" and not match_src and not match_dst and not match_problem:
                                        continue
                                    results.append({
                                        "filename": name,
                                        "index": e.get("index", 0),
                                        "speaker": e.get("name", ""),
                                        "post_src": src_text,
                                        "pre_dst": dst_text,
                                        "match_src": match_src,
                                        "match_dst": match_dst,
                                        "match_problem": match_problem,
                                        "problem": e.get("problem", ""),
                                        "trans_by": e.get("trans_by", ""),
                                    })
                                    if len(results) >= max_results:
                                        break
                            except Exception:
                                continue
                            if len(results) >= max_results:
                                break
                    self._send_json({"results": results, "total": len(results)})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to search cache: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/projects/:id/cache/replace
            if sub_path == "/cache/replace":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    payload = self._read_json_body()
                    query = str(payload.get("query", "")).strip()
                    replacement = str(payload.get("replacement", ""))
                    field = str(payload.get("field", "dst")).strip()  # src | dst | all
                    dry_run = bool(payload.get("dry_run", False))

                    if not query:
                        self._send_json({"error": "empty query"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    import orjson
                    cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                    total_matches = 0
                    total_files = 0
                    file_details = []

                    if os.path.isdir(cache_dir):
                        for name in sorted(os.listdir(cache_dir)):
                            if not name.endswith(".json"):
                                continue
                            fp = os.path.join(cache_dir, name)
                            if not os.path.isfile(fp):
                                continue
                            try:
                                with open(fp, "rb") as f:
                                    entries = orjson.loads(f.read())
                                file_changed = False
                                file_matches = 0
                                for e in entries:
                                    if not isinstance(e, dict):
                                        continue
                                    src_key = "post_src" if "post_src" in e else ("post_jp" if "post_jp" in e else None)
                                    dst_key = "pre_dst" if "pre_dst" in e else ("pre_zh" if "pre_zh" in e else None)
                                    # replace in src
                                    if field in ("src", "all") and src_key and query in e.get(src_key, ""):
                                        if not dry_run:
                                            e[src_key] = e[src_key].replace(query, replacement)
                                        file_matches += 1
                                        file_changed = True
                                    # replace in dst
                                    if field in ("dst", "all") and dst_key and query in e.get(dst_key, ""):
                                        if not dry_run:
                                            e[dst_key] = e[dst_key].replace(query, replacement)
                                        file_matches += 1
                                        file_changed = True
                                    # also replace in proofread_dst / proofread_zh
                                    if field in ("dst", "all"):
                                        pr_key = "proofread_dst" if "proofread_dst" in e else ("proofread_zh" if "proofread_zh" in e else None)
                                        if pr_key and query in e.get(pr_key, ""):
                                            if not dry_run:
                                                e[pr_key] = e[pr_key].replace(query, replacement)
                                            file_matches += 1
                                            file_changed = True
                                if file_matches > 0:
                                    total_matches += file_matches
                                    total_files += 1
                                    detail: dict = {"filename": name, "matches": file_matches}
                                    if file_changed and not dry_run:
                                        detail["entries"] = entries
                                    file_details.append(detail)
                            except Exception:
                                continue
                    self._send_json({
                        "success": True,
                        "total_matches": total_matches,
                        "total_files": total_files,
                        "dry_run": dry_run,
                        "file_details": file_details,
                    })
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to replace in cache: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # GET /api/projects/:id/cache/:filename (catch-all, must be after specific /cache/* routes)
            if sub_path.startswith("/cache/"):
                filename = unquote(sub_path[len("/cache/"):])
                if not filename or filename != os.path.basename(filename):
                    self._send_json({"error": "invalid cache filename"}, status=HTTPStatus.BAD_REQUEST)
                    return
                cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                file_path = os.path.join(cache_dir, filename)
                if not os.path.isfile(file_path):
                    self._send_json({"error": f"cache file not found: {filename}"}, status=HTTPStatus.NOT_FOUND)
                    return
                try:
                    import orjson
                    with open(file_path, "rb") as f:
                        data = orjson.loads(f.read())
                    self._send_json({"project_dir": project_dir, "filename": filename, "entries": data})
                except Exception as exc:
                    self._send_json({"error": f"failed to read cache: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # GET /api/projects/:id/progress
            if sub_path == "/progress":
                cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                total = 0
                translated = 0
                problems = 0
                failed = 0
                file_progress = []
                if os.path.isdir(cache_dir):
                    for name in sorted(os.listdir(cache_dir)):
                        fp = os.path.join(cache_dir, name)
                        if not os.path.isfile(fp) or not fp.endswith(".json"):
                            continue
                        try:
                            import orjson
                            with open(fp, "rb") as f:
                                entries = orjson.loads(f.read())
                            f_total = len(entries)
                            f_translated = sum(1 for e in entries if isinstance(e, dict) and (e.get("pre_dst", "") or e.get("pre_zh", "")))
                            f_problems = sum(1 for e in entries if isinstance(e, dict) and e.get("problem", ""))
                            f_failed = sum(1 for e in entries if isinstance(e, dict) and "(Failed)" in str(e.get("problem", "")))
                            total += f_total
                            translated += f_translated
                            problems += f_problems
                            failed += f_failed
                            file_progress.append({
                                "filename": name,
                                "total": f_total,
                                "translated": f_translated,
                                "problems": f_problems,
                                "failed": f_failed,
                            })
                        except Exception:
                            continue
                self._send_json({
                    "project_dir": project_dir,
                    "total": total,
                    "translated": translated,
                    "problems": problems,
                    "failed": failed,
                    "files": file_progress,
                })
                return

            # GET /api/projects/:id/runtime
            if sub_path == "/runtime":
                runtime = RUNTIME_REGISTRY.get_runtime_snapshot(project_dir)
                file_totals = runtime.get("file_totals", {})
                cache_file_display_map = runtime.get("cache_file_display_map", {})
                config_file_name = "config.yaml"
                job = registry.get_project_job(project_dir)
                if job:
                    config_file_name = job.config_file_name or "config.yaml"
                config_retran_key = RUNTIME_PROGRESS_CACHE.get_retran_key(project_dir, config_file_name)
                retran_terms = _normalize_retran_terms(config_retran_key)
                retran_key = ""
                current_job_started_at_ns = None
                if job and job.status in {"pending", "running"}:
                    retran_key = config_retran_key
                    current_job_started_at_ns = _parse_runtime_job_started_at_ns(job.started_at)
                progress_payload = RUNTIME_PROGRESS_CACHE.get_progress(
                    project_dir,
                    file_totals=file_totals,
                    cache_file_display_map=cache_file_display_map,
                    retran_key=retran_key,
                    retran_terms=retran_terms,
                    current_job_started_at_ns=current_job_started_at_ns,
                )
                total = progress_payload["total"]
                translated = progress_payload["translated"]
                percent = round((translated / total) * 100) if total > 0 else 0
                speed = runtime["translation_speed_lpm"]
                remaining = max(total - translated, 0)
                eta_seconds = round((remaining / speed) * 60) if speed > 0 and remaining > 0 else None
                self._send_json({
                    "project_dir": project_dir,
                    "job": None if job is None else {
                        "job_id": job.job_id,
                        "status": job.status,
                        "translator": job.translator,
                        "created_at": job.created_at,
                        "started_at": job.started_at,
                        "finished_at": job.finished_at,
                        "error": job.error,
                    },
                    "summary": {
                        "total": total,
                        "translated": translated,
                        "problems": progress_payload["problems"],
                        "failed": progress_payload["failed"],
                        "percent": percent,
                        "workers_active": runtime["workers_active"],
                        "workers_configured": runtime["workers_configured"],
                        "translation_speed_lpm": speed,
                        "eta_seconds": eta_seconds,
                        "updated_at": runtime["updated_at"],
                    },
                    "stage": runtime["stage"],
                    "current_file": runtime["current_file"],
                    "recent_errors": runtime["recent_errors"],
                    "recent_successes": runtime["recent_successes"],
                    "retransl_stats": progress_payload["retransl_stats"],
                    "files": progress_payload["files"],
                })
                return

            # POST /api/projects/:id/stop
            if sub_path == "/stop":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                job = registry.request_project_stop(project_dir)
                if job is None:
                    self._send_json(
                        {"success": False, "project_dir": project_dir, "error": "no active job for project"},
                        status=HTTPStatus.CONFLICT,
                    )
                    return
                self._send_json({
                    "success": True,
                    "project_dir": project_dir,
                    "job_id": job.job_id,
                    "status": job.status,
                    "message": "stop requested",
                })
                return

            # GET /api/projects/:id/dictionary
            if sub_path == "/dictionary":
                config_name = parse_qs(urlparse(self.path).query).get("config", ["config.yaml"])[0]
                if not _is_safe_config_filename(config_name):
                    self._send_json({"error": "invalid config filename"}, status=HTTPStatus.BAD_REQUEST)
                    return
                config_path = os.path.join(project_dir, config_name)
                if not os.path.isfile(config_path):
                    self._send_json({"error": f"config file not found: {config_name}"}, status=HTTPStatus.NOT_FOUND)
                    return
                try:
                    data = _read_yaml_file(config_path)
                    dict_cfg = data.get("dictionary", {})
                    default_folder = dict_cfg.get("defaultDictFolder", "Dict")
                    if os.path.isabs(default_folder):
                        dict_base = default_folder
                    else:
                        dict_base = os.path.abspath(default_folder)
                    result = {
                        "project_dir": project_dir,
                        "default_dict_folder": default_folder,
                        "pre_dict_files": dict_cfg.get("preDict", []),
                        "gpt_dict_files": dict_cfg.get("gpt.dict", []),
                        "post_dict_files": dict_cfg.get("postDict", []),
                        "dict_contents": {},
                    }
                    for _, file_list in [
                        ("preDict", dict_cfg.get("preDict", [])),
                        ("gpt.dict", dict_cfg.get("gpt.dict", [])),
                        ("postDict", dict_cfg.get("postDict", [])),
                    ]:
                        for fname in file_list:
                            clean = str(fname).replace(DICT_PROJECT_MARKER, "").strip()
                            if DICT_PROJECT_MARKER in str(fname):
                                fpath = os.path.join(project_dir, clean)
                            else:
                                fpath = os.path.join(dict_base, clean)
                            result["dict_contents"][str(fname)] = _read_dict_file_payload(fpath)
                    self._send_json(result)
                except Exception as exc:
                    self._send_json({"error": f"failed to read dictionary config: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # GET /api/projects/:id/dictionary/project
            if sub_path == "/dictionary/project":
                if self.command != "GET":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                config_name = parse_qs(urlparse(self.path).query).get("config", ["config.yaml"])[0]
                if not _is_safe_config_filename(config_name):
                    self._send_json({"error": "invalid config filename"}, status=HTTPStatus.BAD_REQUEST)
                    return
                try:
                    self._send_json(_collect_project_dict_payload(project_dir, config_name))
                except FileNotFoundError as exc:
                    self._send_json({"error": str(exc)}, status=HTTPStatus.NOT_FOUND)
                except ValueError as exc:
                    self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to load project dictionaries: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/projects/:id/dictionary/project/create
            if sub_path == "/dictionary/project/create":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    payload = self._read_json_body()
                    config_name = str(payload.get("config_file_name", "config.yaml") or "config.yaml")
                    category = str(payload.get("category", "")).strip()
                    filename = str(payload.get("filename", "")).strip()

                    if not _is_safe_config_filename(config_name):
                        self._send_json({"error": "invalid config filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    if not _is_safe_dict_filename(filename):
                        self._send_json({"error": "invalid dictionary filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    config_path = os.path.join(project_dir, config_name)
                    if not os.path.isfile(config_path):
                        self._send_json({"error": f"config file not found: {config_name}"}, status=HTTPStatus.NOT_FOUND)
                        return

                    data = _read_yaml_file(config_path)
                    dict_cfg = data.get("dictionary", {})
                    list_key = _dict_category_config_key(category)
                    current_list = [str(x) for x in dict_cfg.get(list_key, [])]
                    file_key = f"{DICT_PROJECT_MARKER}{filename}"

                    if file_key not in current_list:
                        current_list.append(file_key)
                        dict_cfg[list_key] = current_list
                        data["dictionary"] = dict_cfg
                        _write_yaml_file(config_path, data)

                    file_path = os.path.join(project_dir, filename)
                    os.makedirs(os.path.dirname(file_path) or project_dir, exist_ok=True)
                    if not os.path.exists(file_path):
                        with open(file_path, "w", encoding="utf-8") as f:
                            f.write("")

                    self._send_json({"success": True, "file_key": file_key, "path": file_path})
                except ValueError as exc:
                    self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to create project dictionary file: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/projects/:id/dictionary/project/save
            if sub_path == "/dictionary/project/save":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    payload = self._read_json_body()
                    config_name = str(payload.get("config_file_name", "config.yaml") or "config.yaml")
                    file_key = str(payload.get("file_key", "")).strip()
                    content = _normalize_dict_text(str(payload.get("content", "")))

                    if not _is_safe_config_filename(config_name):
                        self._send_json({"error": "invalid config filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    if DICT_PROJECT_MARKER not in file_key:
                        self._send_json({"error": "file_key must be a project dictionary"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    clean_name = file_key.replace(DICT_PROJECT_MARKER, "").strip()
                    if not _is_safe_dict_filename(clean_name):
                        self._send_json({"error": "invalid dictionary filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    config_path = os.path.join(project_dir, config_name)
                    if not os.path.isfile(config_path):
                        self._send_json({"error": f"config file not found: {config_name}"}, status=HTTPStatus.NOT_FOUND)
                        return

                    data = _read_yaml_file(config_path)
                    dict_cfg = data.get("dictionary", {})
                    listed = set(str(x) for x in dict_cfg.get("preDict", []))
                    listed.update(str(x) for x in dict_cfg.get("gpt.dict", []))
                    listed.update(str(x) for x in dict_cfg.get("postDict", []))
                    if file_key not in listed:
                        self._send_json({"error": "dictionary file is not configured in project dictionary lists"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    file_path = os.path.join(project_dir, clean_name)
                    with open(file_path, "w", encoding="utf-8") as f:
                        f.write(content)
                    self._send_json({"success": True, "file_key": file_key})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to save project dictionary file: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/projects/:id/dictionary/project/delete
            if sub_path == "/dictionary/project/delete":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    payload = self._read_json_body()
                    config_name = str(payload.get("config_file_name", "config.yaml") or "config.yaml")
                    file_key = str(payload.get("file_key", "")).strip()
                    delete_file = bool(payload.get("delete_file", True))

                    if not _is_safe_config_filename(config_name):
                        self._send_json({"error": "invalid config filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    if DICT_PROJECT_MARKER not in file_key:
                        self._send_json({"error": "file_key must be a project dictionary"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    clean_name = file_key.replace(DICT_PROJECT_MARKER, "").strip()
                    if not _is_safe_dict_filename(clean_name):
                        self._send_json({"error": "invalid dictionary filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    config_path = os.path.join(project_dir, config_name)
                    if not os.path.isfile(config_path):
                        self._send_json({"error": f"config file not found: {config_name}"}, status=HTTPStatus.NOT_FOUND)
                        return

                    data = _read_yaml_file(config_path)
                    dict_cfg = data.get("dictionary", {})
                    for list_key in ("preDict", "gpt.dict", "postDict"):
                        current = [str(x) for x in dict_cfg.get(list_key, [])]
                        dict_cfg[list_key] = [x for x in current if x != file_key]
                    data["dictionary"] = dict_cfg
                    _write_yaml_file(config_path, data)

                    if delete_file:
                        file_path = os.path.join(project_dir, clean_name)
                        if os.path.isfile(file_path):
                            os.remove(file_path)

                    self._send_json({"success": True, "file_key": file_key, "deleted_file": delete_file})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to delete project dictionary file: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # GET /api/projects/:id/name-table
            if sub_path == "/name-table":
                # Read the name replacement table (CSV or XLSX)
                csv_path = os.path.join(project_dir, "name替换表.csv")
                xlsx_path = os.path.join(project_dir, "name替换表.xlsx")
                names = []
                source_file = None

                if os.path.isfile(csv_path):
                    source_file = "name替换表.csv"
                    try:
                        import csv as _csv
                        with open(csv_path, "r", newline="", encoding="utf-8-sig") as f:
                            reader = _csv.reader(f)
                            header = next(reader, None)
                            if header:
                                try:
                                    src_idx = header.index("SRC_Name")
                                    dst_idx = header.index("DST_Name")
                                except ValueError:
                                    try:
                                        src_idx = header.index("JP_Name")
                                        dst_idx = header.index("CN_Name")
                                    except ValueError:
                                        self._send_json({"error": "CSV缺少 SRC_Name/DST_Name (或旧版 JP_Name/CN_Name) 列"}, status=HTTPStatus.BAD_REQUEST)
                                        return
                                count_idx = header.index("Count") if "Count" in header else -1
                                for row in reader:
                                    if len(row) > max(src_idx, dst_idx):
                                        names.append({
                                            "src_name": row[src_idx],
                                            "dst_name": row[dst_idx] if dst_idx < len(row) else "",
                                            "count": int(row[count_idx]) if count_idx >= 0 and count_idx < len(row) and row[count_idx].isdigit() else 0,
                                        })
                    except Exception as exc:
                        self._send_json({"error": f"读取CSV人名表失败: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                        return
                elif os.path.isfile(xlsx_path):
                    source_file = "name替换表.xlsx"
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(xlsx_path)
                        sheet = wb.active
                        header = [cell.value for cell in sheet[1]]
                        try:
                            src_idx = header.index("SRC_Name")
                            dst_idx = header.index("DST_Name")
                        except ValueError:
                            try:
                                src_idx = header.index("JP_Name")
                                dst_idx = header.index("CN_Name")
                            except ValueError:
                                self._send_json({"error": "XLSX缺少 SRC_Name/DST_Name (或旧版 JP_Name/CN_Name) 列"}, status=HTTPStatus.BAD_REQUEST)
                                return
                        count_idx = header.index("Count") if "Count" in header else -1
                        for row in sheet.iter_rows(min_row=2):
                            src_val = row[src_idx].value if src_idx < len(row) else None
                            dst_val = row[dst_idx].value if dst_idx < len(row) else None
                            count_val = row[count_idx].value if count_idx >= 0 and count_idx < len(row) else 0
                            if src_val is not None:
                                names.append({
                                    "src_name": str(src_val),
                                    "dst_name": str(dst_val) if dst_val is not None else "",
                                    "count": int(count_val) if count_val else 0,
                                })
                    except Exception as exc:
                        self._send_json({"error": f"读取XLSX人名表失败: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                        return

                self._send_json({
                    "project_dir": project_dir,
                    "source_file": source_file,
                    "names": names,
                })
                return

            # POST /api/projects/:id/name-table/generate
            if sub_path == "/name-table/generate":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                # Generate name table by submitting a dump-name job.
                # This reuses the full pipeline (frontend plugins, splitter, etc.)
                # so speaker names are extracted correctly regardless of input format.
                try:
                    config_name = parse_qs(urlparse(self.path).query).get("config", ["config.yaml"])[0]
                    result = registry.submit({
                        "project_dir": project_dir,
                        "config_file_name": config_name,
                        "translator": "dump-name",
                    })
                    self._send_json({
                        "success": True,
                        "job_id": result.get("job_id", ""),
                    })
                except _ConcurrentLimitError as exc:
                    self._send_json({"error": str(exc)}, status=HTTPStatus.TOO_MANY_REQUESTS)
                except ValueError as exc:
                    self._send_json({"error": str(exc)}, status=HTTPStatus.CONFLICT)
                except Exception as exc:
                    self._send_json({"error": f"生成人名表失败: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/projects/:id/name-table/ai-translate  (SSE streaming)
            if sub_path == "/name-table/ai-translate":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    payload = self._read_json_body()
                    backend_profile = str(payload.get("backend_profile", "")).strip()
                    backend_profile_data = payload.get("backend_profile_data")
                    untranslated = payload.get("names", [])
                    if not isinstance(untranslated, list) or not untranslated:
                        self._send_json({"error": "names must be a non-empty array"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    oai_section = None
                    if isinstance(backend_profile_data, dict) and backend_profile_data:
                        candidate = backend_profile_data.get("OpenAI-Compatible")
                        if isinstance(candidate, dict):
                            oai_section = candidate
                    else:
                        profiles_data = _read_backend_profiles()
                        profiles = profiles_data.get("profiles", {})
                        if backend_profile and backend_profile in profiles:
                            candidate = profiles[backend_profile].get("OpenAI-Compatible")
                            if isinstance(candidate, dict):
                                oai_section = candidate
                        elif backend_profile:
                            self._send_json({"error": f"后端配置 '{backend_profile}' 不存在"}, status=HTTPStatus.NOT_FOUND)
                            return
                        else:
                            for _pname, _pconf in profiles.items():
                                if "OpenAI-Compatible" in _pconf and isinstance(_pconf["OpenAI-Compatible"], dict):
                                    oai_section = _pconf["OpenAI-Compatible"]
                                    break

                    if not oai_section:
                        self._send_json({"error": "未找到可用的 OpenAI 兼容后端配置，请先在后端配置中添加 OpenAI 兼容接口"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    tokens = oai_section.get("tokens", [])
                    if not tokens:
                        self._send_json({"error": "后端配置中没有 API token"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    token_entry = tokens[0]
                    api_key = token_entry.get("token", "")
                    endpoint = token_entry.get("endpoint", "https://api.openai.com")
                    model_name = token_entry.get("modelName", oai_section.get("rewriteModelName", "gpt-4o-mini"))
                    timeout = oai_section.get("apiTimeout", 60)

                    if not api_key or "-example-" in api_key:
                        self._send_json({"error": "后端配置中的 API token 无效"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    import re as _re
                    if endpoint.endswith("/chat/completions"):
                        endpoint = endpoint.replace("/chat/completions", "")
                    if not _re.search(r"/v\d+", endpoint):
                        base_path = "/v1"
                    else:
                        base_path = ""
                    base_url = endpoint.strip("/") + base_path

                    src_names = [str(n.get("src_name", "")) for n in untranslated if str(n.get("src_name", "")).strip()]
                    if not src_names:
                        self._send_json({"error": "没有需要翻译的人名"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    # --- SSE streaming response ---
                    self.send_response(HTTPStatus.OK)
                    self.send_header("Content-Type", "text/event-stream; charset=utf-8")
                    self.send_header("Cache-Control", "no-cache")
                    self.send_header("Connection", "keep-alive")
                    self.end_headers()

                    def _sse_send(event: str, data: dict) -> None:
                        msg = f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"
                        self.wfile.write(msg.encode("utf-8"))
                        self.wfile.flush()

                    names_text = "\n".join(f"- {n}" for n in src_names)
                    system_prompt = (
                        "你是一个专业的日语/中文人名翻译专家。"
                        "请将以下日本人名翻译为中文译名。"
                        "以JSONL格式回答，每行一个JSON对象，格式为："
                        '{"src":"原名","dst":"译名"}\n'
                        "每翻译一个人名就输出一行，不要输出其他任何内容。"
                        "如果某个名字不确定，请尽量给出最常用的中文译名。"
                    )
                    user_prompt = f"请翻译以下人名：\n{names_text}"

                    from openai import OpenAI
                    client = OpenAI(api_key=api_key, base_url=base_url)
                    stream = client.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": user_prompt},
                        ],
                        timeout=timeout,
                        stream=True,
                    )

                    line_buf = ""
                    for chunk in stream:
                        if not chunk.choices:
                            continue
                        delta = chunk.choices[0].delta
                        if not delta or not delta.content:
                            continue
                        line_buf += delta.content
                        # Try to extract complete JSONL lines
                        while "\n" in line_buf:
                            line, line_buf = line_buf.split("\n", 1)
                            line = line.strip()
                            if not line:
                                continue
                            # Strip markdown fences if present
                            if line.startswith("```"):
                                continue
                            try:
                                obj = json.loads(line)
                                src = str(obj.get("src", "")).strip()
                                dst = str(obj.get("dst", "")).strip()
                                if src and dst:
                                    _sse_send("name", {"src_name": src, "dst_name": dst})
                            except (json.JSONDecodeError, ValueError):
                                # Not valid JSON — might be partial, skip
                                pass

                    # Flush remaining buffer
                    if line_buf.strip():
                        line = line_buf.strip()
                        if not line.startswith("```"):
                            try:
                                obj = json.loads(line)
                                src = str(obj.get("src", "")).strip()
                                dst = str(obj.get("dst", "")).strip()
                                if src and dst:
                                    _sse_send("name", {"src_name": src, "dst_name": dst})
                            except (json.JSONDecodeError, ValueError):
                                pass

                    _sse_send("done", {"total": len(untranslated)})

                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    # If headers already sent (SSE started), send error as event
                    try:
                        err_msg = f"AI翻译人名失败: {exc}"
                        self.wfile.write(f"event: error\ndata: {json.dumps({'error': err_msg}, ensure_ascii=False)}\n\n".encode("utf-8"))
                        self.wfile.flush()
                    except Exception:
                        pass
                return

            # POST /api/projects/:id/name-table/save
            if sub_path == "/name-table/save":
                if self.command != "POST":
                    self._send_json({"error": "method not allowed"}, status=HTTPStatus.METHOD_NOT_ALLOWED)
                    return
                try:
                    from GalTransl.Name import write_name_table_csv
                    payload = self._read_json_body()
                    names = payload.get("names", [])
                    if not isinstance(names, list):
                        self._send_json({"error": "names must be an array"}, status=HTTPStatus.BAD_REQUEST)
                        return
                    # Build name_counter and dst_names from the posted data
                    name_counter: dict[str, int] = {}
                    dst_names: dict[str, str] = {}
                    for item in names:
                        src_name = str(item.get("src_name", "") or item.get("jp_name", ""))
                        dst_name = str(item.get("dst_name", "") or item.get("cn_name", ""))
                        count = int(item.get("count", 0))
                        if src_name:
                            name_counter[src_name] = count
                            if dst_name:
                                dst_names[src_name] = dst_name
                    csv_path = os.path.join(project_dir, "name替换表.csv")
                    write_name_table_csv(csv_path, name_counter, dst_names)
                    self._send_json({"success": True, "source_file": "name替换表.csv", "total": len(names)})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"保存人名表失败: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # GET /api/projects/:id/name-dict
            if sub_path == "/name-dict":
                # Return the name replacement dict (SRC→DST) for UI pill display
                csv_path = os.path.join(project_dir, "name替换表.csv")
                xlsx_path = os.path.join(project_dir, "name替换表.xlsx")
                name_dict: dict[str, str] = {}

                def _find_col(header, new_name, old_name):
                    """Find column index, preferring new name, falling back to old."""
                    if new_name in header:
                        return header.index(new_name)
                    if old_name in header:
                        return header.index(old_name)
                    return -1

                if os.path.isfile(csv_path):
                    try:
                        import csv as _csv
                        with open(csv_path, "r", newline="", encoding="utf-8-sig") as f:
                            reader = _csv.reader(f)
                            header = next(reader, None)
                            if header:
                                src_idx = _find_col(header, "SRC_Name", "JP_Name")
                                dst_idx = _find_col(header, "DST_Name", "CN_Name")
                                if src_idx >= 0 and dst_idx >= 0:
                                    for row in reader:
                                        if len(row) > max(src_idx, dst_idx):
                                            dst = row[dst_idx].strip() if dst_idx < len(row) else ""
                                            if dst:
                                                name_dict[row[src_idx]] = dst
                    except Exception:
                        pass
                elif os.path.isfile(xlsx_path):
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(xlsx_path)
                        sheet = wb.active
                        header = [cell.value for cell in sheet[1]]
                        src_idx = _find_col(header, "SRC_Name", "JP_Name")
                        dst_idx = _find_col(header, "DST_Name", "CN_Name")
                        if src_idx >= 0 and dst_idx >= 0:
                            for row in sheet.iter_rows(min_row=2):
                                src_val = row[src_idx].value if src_idx < len(row) else None
                                dst_val = row[dst_idx].value if dst_idx < len(row) else None
                                if src_val is not None and dst_val is not None and str(dst_val).strip():
                                    name_dict[str(src_val)] = str(dst_val)
                    except Exception:
                        pass

                self._send_json({"project_dir": project_dir, "name_dict": name_dict})
                return

            # GET /api/projects/:id/problems
            if sub_path == "/problems":
                cache_dir = os.path.join(project_dir, CACHE_FOLDERNAME)
                all_problems = []
                if os.path.isdir(cache_dir):
                    for name in sorted(os.listdir(cache_dir)):
                        fp = os.path.join(cache_dir, name)
                        if not os.path.isfile(fp) or not fp.endswith(".json"):
                            continue
                        try:
                            import orjson
                            with open(fp, "rb") as f:
                                entries = orjson.loads(f.read())
                            for e in entries:
                                if isinstance(e, dict) and e.get("problem", ""):
                                    all_problems.append({
                                        "filename": name,
                                        "index": e.get("index", 0),
                                        "speaker": e.get("name", ""),
                                        "post_src": e.get("post_src", "") or e.get("post_jp", ""),
                                        "pre_dst": e.get("pre_dst", "") or e.get("pre_zh", ""),
                                        "problem": e.get("problem", ""),
                                        "trans_by": e.get("trans_by", ""),
                                    })
                        except Exception:
                            continue
                self._send_json({"project_dir": project_dir, "problems": all_problems, "total": len(all_problems)})
                return

            # GET /api/projects/:id/logs
            if sub_path == "/logs":
                log_path = os.path.join(project_dir, "GalTransl.log")
                if not os.path.isfile(log_path):
                    self._send_json({"project_dir": project_dir, "exists": False, "lines": []})
                    return
                try:
                    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
                        lines = f.read().splitlines()
                    # Return last 2000 lines by default
                    query = parse_qs(urlparse(self.path).query)
                    tail = int(query.get("tail", ["2000"])[0])
                    self._send_json({
                        "project_dir": project_dir,
                        "exists": True,
                        "total_lines": len(lines),
                        "lines": lines[-tail:],
                    })
                except Exception as exc:
                    self._send_json({"error": f"failed to read log: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            self._send_json({"error": "not found"}, status=HTTPStatus.NOT_FOUND)

        def do_GET(self) -> None:
            parsed = urlparse(self.path)
            path = parsed.path

            if path == "/":
                self._send_html(INDEX_HTML)
                return
            if path == "/api/version":
                self._send_json({"version": GALTRANSL_VERSION})
                return
            if path == "/api/version/check":
                latest_version = new_version[0] if new_version else None
                self._send_json(
                    {
                        "version": GALTRANSL_VERSION,
                        "latest_version": latest_version,
                        "update_available": _has_newer_release(GALTRANSL_VERSION, latest_version),
                    }
                )
                return
            if path == "/api/translators":
                _hidden_translators = {"show-plugs", "dump-name", "rebuildr", "rebuilda"}
                translators = [
                    {
                        "name": name,
                        "description": description.get("zh-cn") or next(iter(description.values())),
                    }
                    for name, description in TRANSLATOR_SUPPORTED.items()
                    if name not in _hidden_translators
                ]
                self._send_json({"translators": translators})
                return
            if path == "/api/jobs":
                self._send_json({"jobs": registry.list_jobs()})
                return
            if path == "/api/app-settings":
                self._send_json(load_app_settings())
                return
            if path == "/api/project-config-template":
                self._send_json({"content": DEFAULT_PROJECT_CONFIG_YAML})
                return
            if path.startswith("/api/jobs/"):
                job_id = path.rsplit("/", 1)[-1]
                job = registry.get_job(job_id)
                if job is None:
                    self._send_json({"error": "job not found"}, status=HTTPStatus.NOT_FOUND)
                    return
                self._send_json(job)
                return

            # --- Backend Profiles API ---
            if path == "/api/backend-profiles":
                data = _read_backend_profiles()
                self._send_json(data)
                return

            if path.startswith("/api/backend-profiles/"):
                profile_name = unquote(path.split("/", 3)[-1])
                data = _read_backend_profiles()
                profiles = data.get("profiles", {})
                if profile_name not in profiles:
                    self._send_json({"error": f"profile not found: {profile_name}"}, status=HTTPStatus.NOT_FOUND)
                    return
                self._send_json({"name": profile_name, "profile": profiles[profile_name]})
                return

            # --- Project API ---
            if path == "/api/plugins":
                self._send_json({"plugins": _scan_plugins()})
                return

            if path == "/api/problem-types":
                self._send_json({"problem_types": _list_problem_types()})
                return

            if path == "/api/translation-guidelines":
                self._send_json({"guidelines": _list_translation_guidelines()})
                return

            if path.startswith("/api/projects/"):
                parts = path.split("/", 4)  # /api/projects/:id/sub...
                if len(parts) < 4:
                    self._send_json({"error": "invalid project path"}, status=HTTPStatus.BAD_REQUEST)
                    return
                project_id = parts[3]
                sub_path = "/" + "/".join(parts[4:]) if len(parts) > 4 else "/"
                self._route_project_api(project_id, sub_path)
                return

            # GET /api/dictionaries/common
            if path == "/api/dictionaries/common":
                try:
                    self._send_json(_collect_common_dict_payload())
                except Exception as exc:
                    self._send_json({"error": f"failed to load common dictionaries: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            self._send_json({"error": "not found"}, status=HTTPStatus.NOT_FOUND)

        def do_POST(self) -> None:
            parsed = urlparse(self.path)
            path = parsed.path

            if path.startswith("/api/projects/"):
                parts = path.split("/", 4)
                if len(parts) < 4:
                    self._send_json({"error": "invalid project path"}, status=HTTPStatus.BAD_REQUEST)
                    return
                project_id = parts[3]
                sub_path = "/" + "/".join(parts[4:]) if len(parts) > 4 else "/"
                self._route_project_api(project_id, sub_path)
                return

            # POST /api/openai-models — query a list of models from an OpenAI-compatible API.
            if path == "/api/openai-models":
                try:
                    payload = self._read_json_body()
                    endpoint = str(payload.get("endpoint", "")).strip() or "https://api.openai.com"
                    token = str(payload.get("token", "")).strip()
                    proxy = payload.get("proxy")
                    timeout = float(payload.get("timeout", 15))

                    base = endpoint.rstrip("/")
                    if base.endswith("/v1"):
                        url = base + "/models"
                    elif base.rstrip("/").endswith("/models"):
                        url = base
                    else:
                        url = base + "/v1/models"

                    import urllib.request
                    import urllib.error
                    req = urllib.request.Request(url, method="GET")
                    if token:
                        req.add_header("Authorization", f"Bearer {token}")
                    req.add_header("Accept", "application/json")

                    opener_args = []
                    if isinstance(proxy, dict):
                        proxy_map: dict[str, str] = {}
                        http_proxy = str(proxy.get("http") or proxy.get("http_proxy") or "").strip()
                        https_proxy = str(proxy.get("https") or proxy.get("https_proxy") or http_proxy).strip()
                        if http_proxy:
                            proxy_map["http"] = http_proxy
                        if https_proxy:
                            proxy_map["https"] = https_proxy
                        if proxy_map:
                            opener_args.append(urllib.request.ProxyHandler(proxy_map))
                    elif isinstance(proxy, str) and proxy.strip():
                        opener_args.append(urllib.request.ProxyHandler({"http": proxy.strip(), "https": proxy.strip()}))
                    else:
                        # Explicitly bypass system proxies when none provided to avoid unexpected routing.
                        opener_args.append(urllib.request.ProxyHandler({}))

                    opener = urllib.request.build_opener(*opener_args)
                    try:
                        with opener.open(req, timeout=timeout) as resp:
                            raw = resp.read()
                    except urllib.error.HTTPError as http_exc:
                        try:
                            body = http_exc.read().decode("utf-8", errors="replace")
                        except Exception:
                            body = ""
                        self._send_json(
                            {"error": f"HTTP {http_exc.code}: {http_exc.reason}", "detail": body[:1000]},
                            status=HTTPStatus.BAD_GATEWAY,
                        )
                        return
                    except urllib.error.URLError as url_exc:
                        self._send_json({"error": f"请求失败: {url_exc.reason}"}, status=HTTPStatus.BAD_GATEWAY)
                        return

                    try:
                        data = json.loads(raw.decode("utf-8", errors="replace"))
                    except json.JSONDecodeError:
                        self._send_json({"error": "响应不是合法的 JSON"}, status=HTTPStatus.BAD_GATEWAY)
                        return

                    # OpenAI-style: {"data": [{"id": "..."}, ...]}
                    # Some backends may return a bare list.
                    items: list[Any] = []
                    if isinstance(data, dict):
                        if isinstance(data.get("data"), list):
                            items = data["data"]
                        elif isinstance(data.get("models"), list):
                            items = data["models"]
                    elif isinstance(data, list):
                        items = data

                    models: list[str] = []
                    for item in items:
                        if isinstance(item, str):
                            models.append(item)
                        elif isinstance(item, dict):
                            mid = item.get("id") or item.get("name") or item.get("model")
                            if isinstance(mid, str) and mid:
                                models.append(mid)
                    # De-duplicate while preserving order.
                    seen = set()
                    unique_models: list[str] = []
                    for m in models:
                        if m not in seen:
                            seen.add(m)
                            unique_models.append(m)

                    self._send_json({"models": unique_models, "url": url})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to fetch models: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            if path == "/api/jobs":
                try:
                    payload = self._read_json_body()
                    job = registry.submit(payload)
                except _ConcurrentLimitError as exc:
                    self._send_json({"error": str(exc)}, status=HTTPStatus.TOO_MANY_REQUESTS)
                    return
                except ValueError as exc:
                    self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
                    return
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                    return

                self._send_json(job, status=HTTPStatus.ACCEPTED)
                return

            # POST /api/dictionaries/common/create
            if path == "/api/dictionaries/common/create":
                try:
                    payload = self._read_json_body()
                    category = str(payload.get("category", "")).strip()
                    filename = str(payload.get("filename", "")).strip()

                    _dict_category_config_key(category)
                    if not _is_safe_dict_filename(filename):
                        self._send_json({"error": "invalid dictionary filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    dict_dir = _common_dict_directory()
                    os.makedirs(dict_dir, exist_ok=True)
                    file_path = os.path.join(dict_dir, filename)
                    if not os.path.exists(file_path):
                        with open(file_path, "w", encoding="utf-8") as f:
                            f.write("")

                    category_map = _read_common_dict_category_map(dict_dir)
                    category_map[filename] = category
                    _write_common_dict_category_map(dict_dir, category_map)

                    self._send_json({"success": True, "filename": filename, "path": file_path})
                except ValueError as exc:
                    self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to create common dictionary file: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/dictionaries/common/save
            if path == "/api/dictionaries/common/save":
                try:
                    payload = self._read_json_body()
                    filename = str(payload.get("filename", "")).strip()
                    content = _normalize_dict_text(str(payload.get("content", "")))

                    if not _is_safe_dict_filename(filename):
                        self._send_json({"error": "invalid dictionary filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    dict_dir = _common_dict_directory()
                    os.makedirs(dict_dir, exist_ok=True)
                    file_path = os.path.join(dict_dir, filename)
                    with open(file_path, "w", encoding="utf-8") as f:
                        f.write(content)

                    self._send_json({"success": True, "filename": filename})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to save common dictionary file: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # POST /api/dictionaries/common/delete
            if path == "/api/dictionaries/common/delete":
                try:
                    payload = self._read_json_body()
                    filename = str(payload.get("filename", "")).strip()
                    if not _is_safe_dict_filename(filename):
                        self._send_json({"error": "invalid dictionary filename"}, status=HTTPStatus.BAD_REQUEST)
                        return

                    file_path = os.path.join(_common_dict_directory(), filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)

                    dict_dir = _common_dict_directory()
                    category_map = _read_common_dict_category_map(dict_dir)
                    if filename in category_map:
                        del category_map[filename]
                        _write_common_dict_category_map(dict_dir, category_map)

                    self._send_json({"success": True, "filename": filename})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to delete common dictionary file: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            self._send_json({"error": "not found"}, status=HTTPStatus.NOT_FOUND)

        def do_PUT(self) -> None:
            parsed = urlparse(self.path)
            path = parsed.path

            if path == "/api/app-settings":
                try:
                    payload = self._read_json_body()
                    settings = save_app_settings(payload)
                    self._send_json(settings)
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to write app settings: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # PUT /api/backend-profiles/:name
            if path.startswith("/api/backend-profiles/"):
                profile_name = unquote(path.split("/", 3)[-1])
                if not profile_name:
                    self._send_json({"error": "profile name is required"}, status=HTTPStatus.BAD_REQUEST)
                    return
                try:
                    payload = self._read_json_body()
                    profile_data = payload.get("profile")
                    if profile_data is None:
                        self._send_json({"error": "profile field is required"}, status=HTTPStatus.BAD_REQUEST)
                        return
                    data = _read_backend_profiles()
                    if "profiles" not in data:
                        data["profiles"] = {}
                    data["profiles"][profile_name] = profile_data
                    _write_backend_profiles(data)
                    self._send_json({"success": True, "name": profile_name})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to write profile: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            # PUT /api/projects/:id/config
            if path.startswith("/api/projects/") and path.endswith("/config"):
                parts = path.split("/")
                if len(parts) < 5:
                    self._send_json({"error": "invalid project path"}, status=HTTPStatus.BAD_REQUEST)
                    return
                project_id = parts[3]
                try:
                    project_dir = _safe_project_dir(project_id)
                except ValueError as exc:
                    self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
                    return
                try:
                    payload = self._read_json_body()
                    config_data = payload.get("config")
                    config_name = payload.get("config_file_name", "config.yaml")
                    if config_data is None:
                        self._send_json({"error": "config field is required"}, status=HTTPStatus.BAD_REQUEST)
                        return
                    config_path = os.path.join(project_dir, config_name)
                    if not os.path.isfile(config_path):
                        self._send_json({"error": f"config file not found: {config_name}"}, status=HTTPStatus.NOT_FOUND)
                        return
                    _write_yaml_file(config_path, config_data)
                    self._send_json({"success": True, "project_dir": project_dir, "config_file_name": config_name})
                except json.JSONDecodeError:
                    self._send_json({"error": "invalid json body"}, status=HTTPStatus.BAD_REQUEST)
                except Exception as exc:
                    self._send_json({"error": f"failed to write config: {exc}"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return

            self._send_json({"error": "not found"}, status=HTTPStatus.NOT_FOUND)

        def do_DELETE(self) -> None:
            parsed = urlparse(self.path)
            path = parsed.path

            # DELETE /api/backend-profiles/:name
            if path.startswith("/api/backend-profiles/"):
                profile_name = unquote(path.split("/", 3)[-1])
                if not profile_name:
                    self._send_json({"error": "profile name is required"}, status=HTTPStatus.BAD_REQUEST)
                    return
                data = _read_backend_profiles()
                profiles = data.get("profiles", {})
                if profile_name not in profiles:
                    self._send_json({"error": f"profile not found: {profile_name}"}, status=HTTPStatus.NOT_FOUND)
                    return
                del data["profiles"][profile_name]
                _write_backend_profiles(data)
                self._send_json({"success": True, "name": profile_name})
                return

            self._send_json({"error": "not found"}, status=HTTPStatus.NOT_FOUND)

        def log_message(self, format: str, *args: Any) -> None:
            return

        def _read_json_body(self) -> dict[str, Any]:
            content_length = int(self.headers.get("Content-Length", "0"))
            raw_body = self.rfile.read(content_length) if content_length > 0 else b"{}"
            data = json.loads(raw_body.decode("utf-8") or "{}")
            if not isinstance(data, dict):
                raise ValueError("json body must be an object")
            return data

        def _send_json(self, payload: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _send_html(self, html: str) -> None:
            body = html.encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

    return RequestHandler


def serve(host: str = "127.0.0.1", port: int = 12333) -> None:
    registry = JobRegistry()
    try:
        server = ThreadingHTTPServer((host, port), build_handler(registry))
    except OSError as exc:
        # WinError 10048 / errno 98 (EADDRINUSE) / errno 13 (EACCES on Windows for occupied ports)
        errno_val = getattr(exc, "errno", None)
        winerror = getattr(exc, "winerror", None)
        if errno_val in (48, 98, 10048, 13) or winerror == 10048:
            print(
                f"[错误] 端口 {port} 已被占用，无法启动 GalTransl 后端服务。\n"
                f"       请先关闭占用该端口的程序，或使用 --port 指定其他端口，例如：\n"
                f"       python run_backend.py --host {host} --port {port + 1}"
            )
            raise SystemExit(1)
        print(f"[错误] 无法绑定 {host}:{port} —— {exc}")
        raise SystemExit(1)
    print(f"GalTransl backend mode listening at http://{host}:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


def main() -> None:
    parser = argparse.ArgumentParser("GalTransl backend mode")
    parser.add_argument("--host", default="127.0.0.1", help="bind host")
    parser.add_argument("--port", type=int, default=12333, help="bind port")
    args = parser.parse_args()
    serve(args.host, args.port)


if __name__ == "__main__":
    main()
